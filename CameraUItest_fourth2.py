import os
import subprocess
import sys
import shutil
import re
import numpy as np
import json
import time
from typing import Union
from datetime import datetime

from mecheye.shared import *
from mecheye.area_scan_3d_camera import *
from mecheye.area_scan_3d_camera_utils import *

try:
    import open3d as o3d
except ImportError:
    o3d = None

# pyqtgraph 3D 뷰어
try:
    import pyqtgraph as pg
    from pyqtgraph.opengl import GLViewWidget, GLScatterPlotItem
except ImportError:
    pg = None
    GLViewWidget = None
    GLScatterPlotItem = None

# 깊이 TIFF / 이미지 읽기용
try:
    import cv2
except ImportError:
    cv2 = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font
except ImportError:
    openpyxl = None

from PySide6.QtCore import Qt, QObject, Signal, QThread, QEvent, QTimer, QPoint, QAbstractTableModel, QModelIndex, QSize
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget,
    QVBoxLayout, QHBoxLayout, QPushButton,
    QTextEdit, QLabel, QSizePolicy, QButtonGroup, QMessageBox,
    QDialog, QListWidget, QLineEdit, QCheckBox,
    QGridLayout, QFrame, QTableView, QHeaderView,
    QToolBar, QDockWidget, QGroupBox, QFormLayout, QSplitter, QFileDialog,
    QSplitterHandle,  # ✅ 추가
)
from PySide6.QtGui import QPalette, QColor, QPixmap, QImage, QGuiApplication, QCursor

# --------------------------------------------------------------------
# ★ 실제 카메라를 쓸 때는 True 로 바꾸기.
# --------------------------------------------------------------------
USE_REAL_CAMERA = False

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

SAMPLE_PLY_PATH = os.path.join(BASE_DIR, "sample", "sample_point.ply")
SAMPLE_TIFF_PATH = os.path.join(BASE_DIR, "sample", "sample_depth.tiff")
SAMPLE_IMG_PATH = os.path.join(BASE_DIR, "sample", "sampleimg.png")

# 결과 저장 폴더(최근 스캔 저장)
RESULT_DIR = os.path.join(BASE_DIR, "result", "objects_ply_data", "2025")

# ★ 비교데이터 저장 폴더(요청사항)
COMPARE_DIR = os.path.join(BASE_DIR, "comparedata")

ERROR_LOG_DIR = os.path.join(BASE_DIR, "error_log")   # .\error_log
ERROR_SAVE_DIR = os.path.join(BASE_DIR, "error")      # .\error

# ----------------- 깊이값 → 레인보우 컬러맵 -----------------
def hsv_to_rgb_np(h, s, v):
    """h,s,v in [0,1] numpy 배열 → r,g,b in [0,1] numpy 배열."""
    h = np.mod(h, 1.0)
    i = np.floor(h * 6).astype(int)
    f = h * 6 - i

    p = v * (1.0 - s)
    q = v * (1.0 - f * s)
    t = v * (1.0 - (1.0 - f) * s)

    i_mod = i % 6

    r = np.zeros_like(h)
    g = np.zeros_like(h)
    b = np.zeros_like(h)

    mask = (i_mod == 0)
    r[mask], g[mask], b[mask] = v[mask], t[mask], p[mask]

    mask = (i_mod == 1)
    r[mask], g[mask], b[mask] = q[mask], v[mask], p[mask]

    mask = (i_mod == 2)
    r[mask], g[mask], b[mask] = p[mask], v[mask], t[mask]

    mask = (i_mod == 3)
    r[mask], g[mask], b[mask] = p[mask], q[mask], v[mask]

    mask = (i_mod == 4)
    r[mask], g[mask], b[mask] = t[mask], p[mask], v[mask]

    mask = (i_mod == 5)
    r[mask], g[mask], b[mask] = v[mask], p[mask], q[mask]

    return r, g, b


def depth_to_color_image(depth_np: np.ndarray) -> np.ndarray:
    """
    depth_np: 2D float32 배열 (깊이값)
    return: HxWx3 uint8 RGB 컬러맵 이미지
    """
    depth = depth_np.astype(np.float32)

    depth = np.nan_to_num(depth, nan=0.0, posinf=0.0, neginf=0.0)

    if depth.ndim == 3:
        depth = depth[..., 0]

    mask = np.isfinite(depth) & (depth > 0)
    if not np.any(mask):
        h, w = depth.shape
        return np.zeros((h, w, 3), np.uint8)

    d_min = float(depth[mask].min())
    d_max = float(depth[mask].max())

    if d_max <= d_min:
        h, w = depth.shape
        return np.zeros((h, w, 3), np.uint8)

    norm = (depth - d_min) / (d_max - d_min + 1e-6)
    norm = np.clip(norm, 0.0, 1.0)

    # 가까운 곳(값 작음) = 빨강, 먼 곳(값 큼) = 파랑
    h = (norm) * (2.0 / 3.0)
    s = np.ones_like(h)
    v = np.ones_like(h)

    r, g, b = hsv_to_rgb_np(h, s, v)
    rgb = np.stack([r, g, b], axis=-1)

    rgb_uint8 = (rgb * 255.0).astype(np.uint8)
    rgb_uint8[~mask] = 0
    return rgb_uint8


def sanitize_folder_component(text: str) -> str:
    """폴더/파일명에 위험한 문자 제거."""
    if text is None:
        return ""
    text = text.strip()
    text = re.sub(r'[\\/:*?"<>|]', "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def list_compare_folders() -> list[str]:
    """COMPARE_DIR 안의 폴더 목록을 반환(정렬)."""
    if not os.path.isdir(COMPARE_DIR):
        return []
    items = []
    for name in os.listdir(COMPARE_DIR):
        p = os.path.join(COMPARE_DIR, name)
        if os.path.isdir(p):
            items.append(name)
    items.sort()
    return items


def find_folder_by_number(num_text: str) -> str | None:
    """번호로 시작하는 폴더(예: 12_XXX) 찾기."""
    num_text = (num_text or "").strip()
    if num_text == "":
        return None
    folders = list_compare_folders()
    prefix = f"{num_text}_"
    for f in folders:
        if f.startswith(prefix):
            return f
    return None


# ============================================================
# Hangul Composer (초/중/종성 조합) + Virtual Keyboard
# ============================================================

CHO = list("ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ")
JUNG = list("ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ")
JONG = [""] + list("ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ")

# 복합 모음/받침 규칙
JUNG_COMBINE = {
    ("ㅗ", "ㅏ"): "ㅘ",
    ("ㅗ", "ㅐ"): "ㅙ",
    ("ㅗ", "ㅣ"): "ㅚ",
    ("ㅜ", "ㅓ"): "ㅝ",
    ("ㅜ", "ㅔ"): "ㅞ",
    ("ㅜ", "ㅣ"): "ㅟ",
    ("ㅡ", "ㅣ"): "ㅢ",
}
JUNG_SPLIT = {v: k for k, v in JUNG_COMBINE.items()}

JONG_COMBINE = {
    ("ㄱ", "ㅅ"): "ㄳ",
    ("ㄴ", "ㅈ"): "ㄵ",
    ("ㄴ", "ㅎ"): "ㄶ",
    ("ㄹ", "ㄱ"): "ㄺ",
    ("ㄹ", "ㅁ"): "ㄻ",
    ("ㄹ", "ㅂ"): "ㄼ",
    ("ㄹ", "ㅅ"): "ㄽ",
    ("ㄹ", "ㅌ"): "ㄾ",
    ("ㄹ", "ㅍ"): "ㄿ",
    ("ㄹ", "ㅎ"): "ㅀ",
    ("ㅂ", "ㅅ"): "ㅄ",
}
JONG_SPLIT = {v: k for k, v in JONG_COMBINE.items()}


def is_ja(c: str) -> bool:
    return c in CHO or c in JONG[1:]  # 종성 자모 포함


def is_mo(c: str) -> bool:
    return c in JUNG


def compose_syllable(cho: str, jung: str, jong: str = "") -> str:
    if cho not in CHO or jung not in JUNG:
        return ""
    ci = CHO.index(cho)
    ji = JUNG.index(jung)
    ti = JONG.index(jong) if jong in JONG else 0
    code = 0xAC00 + (ci * 21 * 28) + (ji * 28) + ti
    return chr(code)


def build_view_panel(title: str, viewer: QWidget, buttons: list) -> QWidget:
    root = QGroupBox()
    v = QVBoxLayout(root)
    v.setContentsMargins(8, 8, 8, 8)
    v.setSpacing(6)

    # 헤더(제목 + 버튼들)
    header = QHBoxLayout()
    header.setContentsMargins(0, 0, 0, 0)

    lbl = QLabel(title)
    lbl.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)

    header.addWidget(lbl, 1)  # 제목은 왼쪽에서 늘어남
    for b in buttons:
        header.addWidget(b, 0)  # 버튼은 오른쪽에 붙음

    v.addLayout(header)
    v.addWidget(viewer, 1)
    return root


class FixedSplitterHandle(QSplitterHandle):
    def mousePressEvent(self, e): e.ignore()
    def mouseMoveEvent(self, e): e.ignore()
    def mouseReleaseEvent(self, e): e.ignore()


class FixedSplitter(QSplitter):
    def createHandle(self):
        return FixedSplitterHandle(self.orientation(), self)


class HangulComposer:
    """
    아주 기본적인 한글 조합 상태 머신.
    - 조합중인 글자(초/중/종)를 상태로 갖고 있다가 commit_text / flush로 확정한다.
    """
    def __init__(self):
        self.reset()

    def reset(self):
        self.cho = ""
        self.jung = ""
        self.jong = ""

    def has_composing(self) -> bool:
        return self.cho != "" or self.jung != "" or self.jong != ""

    def composing_text(self) -> str:
        if self.cho and self.jung:
            return compose_syllable(self.cho, self.jung, self.jong)
        # 초성만 있는 상태면 자모 그대로 보이게
        if self.cho and not self.jung:
            return self.cho
        # 모음만 단독 입력되는 경우
        if self.jung and not self.cho:
            return self.jung
        return ""

    def input_char(self, c: str):
        """
        입력 문자 c를 받아서,
        - (commit_string, composing_string) 형태로 반환
        commit_string은 즉시 확정해서 QLineEdit에 넣을 텍스트,
        composing_string은 조합중으로 보일 텍스트(마지막에 덮어쓰기용)
        """
        commit = ""

        # 공백/기타는 조합 확정 후 그대로 커밋
        if not (is_ja(c) or is_mo(c)):
            if self.has_composing():
                commit += self.composing_text()
                self.reset()
            commit += c
            return commit, ""

        # --------------------
        # 모음 입력
        # --------------------
        if is_mo(c):
            # (1) 초성 없이 모음부터 들어오면: 모음 단독
            if self.cho == "" and self.jung == "":
                self.jung = c
                return "", self.composing_text()

            # (2) 초성만 있는 상태 -> 중성 채우기
            if self.cho != "" and self.jung == "":
                self.jung = c
                return "", self.composing_text()

            # (3) 중성이 이미 있는데 모음이 또 오면 복합모음 시도
            if self.jung != "":
                # 종성이 있으면: 종성을 다음 글자 초성으로 넘기는 케이스 처리
                if self.jong != "":
                    # 종성이 복합받침이면 분리
                    if self.jong in JONG_SPLIT:
                        a, b = JONG_SPLIT[self.jong]
                        # 앞 글자에는 a만 종성으로 남기고, b는 다음 글자 초성으로
                        prev = compose_syllable(self.cho, self.jung, a)
                        self.cho = b
                        self.jung = c
                        self.jong = ""
                        commit += prev
                        return commit, self.composing_text()
                    else:
                        # ✅ 단일 종성: 앞 글자는 종성 없이 확정하고,
                        # ✅ 종성을 다음 글자 초성으로 넘김
                        prev = compose_syllable(self.cho, self.jung, "")
                        commit += prev
                        self.cho = self.jong
                        self.jung = c
                        self.jong = ""
                        return commit, self.composing_text()

                # 종성 없으면 복합모음 결합
                key = (self.jung, c)
                if key in JUNG_COMBINE:
                    self.jung = JUNG_COMBINE[key]
                    return "", self.composing_text()
                else:
                    # 복합모음 불가 -> 이전 글자 확정 후 새 모음 시작
                    commit += self.composing_text()
                    self.reset()
                    self.jung = c
                    return commit, self.composing_text()

        # --------------------
        # 자음 입력
        # --------------------
        if is_ja(c):
            # (1) 아무것도 없으면 초성으로 시작
            if self.cho == "" and self.jung == "":
                self.cho = c
                return "", self.composing_text()

            # (2) 초성만 있는데 자음 또 오면: 된소리/겹자음 초성은 여기선 단순 처리(이전 확정 후 새 초성)
            if self.cho != "" and self.jung == "":
                commit += self.cho
                self.cho = c
                return commit, self.composing_text()

            # (3) 초+중 있는 상태에서 자음: 종성으로 들어감
            if self.cho != "" and self.jung != "":
                # 종성이 비어있으면 바로 종성으로
                if self.jong == "":
                    self.jong = c
                    return "", self.composing_text()

                # 종성이 이미 있으면 복합받침 결합 시도
                key = (self.jong, c)
                if key in JONG_COMBINE:
                    self.jong = JONG_COMBINE[key]
                    return "", self.composing_text()

                # 결합 불가: 이전 글자 확정 후 새 초성 시작
                commit += self.composing_text()
                self.reset()
                self.cho = c
                return commit, self.composing_text()

        return "", self.composing_text()

    def backspace(self):
        """
        조합중이면 조합 상태를 한 단계 되돌린다.
        반환: (commit_remove_last_char: bool, new_composing_string)
        - commit_remove_last_char=True면 QLineEdit에서 실제 문자 하나 삭제 필요
        """
        if not self.has_composing():
            return True, ""  # 조합 없으면 실제 문자 삭제

        # 조합 상태 되돌리기: 종 -> 중 -> 초
        if self.jong:
            # 복합받침이면 분리해서 첫번째만 남김
            if self.jong in JONG_SPLIT:
                a, b = JONG_SPLIT[self.jong]
                self.jong = a
            else:
                self.jong = ""
            return False, self.composing_text()

        if self.jung:
            # 복합모음이면 분리해서 첫번째만 남김
            if self.jung in JUNG_SPLIT:
                a, b = JUNG_SPLIT[self.jung]
                self.jung = a
            else:
                self.jung = ""
            return False, self.composing_text()

        if self.cho:
            self.cho = ""
            return False, ""

        return True, ""


class VirtualKeyboard(QWidget):
    """
    mode:
      - "num" : 숫자 키패드(오차/게인/노출/번호검색 등)
      - "full": 전체 키보드(비교데이터 이름 입력 등) + 한/영 토글(자체 조합)
    """
    requestHide = Signal()
    requestReposition = Signal()
    

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setAttribute(Qt.WA_ShowWithoutActivating, True)
        self.setFocusPolicy(Qt.NoFocus)
        self.setWindowModality(Qt.NonModal)
        self.setAttribute(Qt.WA_AcceptTouchEvents, True)

        self.target: QLineEdit | None = None
        self.mode = "num"
        self.lang = "KO"   # "KO" or "EN"
        self.shift = False

        self.composer = HangulComposer()
        self._composing_len = 0  # QLineEdit에 덮어쓴 조합 문자열 길이

        self.root = QVBoxLayout(self)
        self.root.setContentsMargins(8, 8, 8, 8)
        self.root.setSpacing(6)

        self.title = QLabel("Virtual Keyboard")
        self.title.setStyleSheet("color:white; font-size:14px;")
        self.root.addWidget(self.title)

        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setStyleSheet("color:#999;")
        self.root.addWidget(line)

        self.grid = QGridLayout()
        self.grid.setSpacing(6)
        self.root.addLayout(self.grid)

        self.setStyleSheet("""
            QWidget { background-color:#2b2b2b; border:1px solid #555; }
            QPushButton {
                background-color:#3f4c55; color:white;
                padding:10px; font-size:16px; border-radius:6px;
            }
            QPushButton:pressed { background-color:#0e5a7a; }
        """)

        self._rebuild()

    def attach(self, target: QLineEdit, mode: str):
        self.target = target
        self.mode = mode
        self.shift = False
        # 이름 입력은 한글 가능성이 높아서 KO 기본
        self.lang = "KO" if mode == "full" else "EN"
        self.composer.reset()
        self._composing_len = 0
        self._rebuild()

    def _clear_grid(self):
        while self.grid.count():
            item = self.grid.takeAt(0)
            w = item.widget()
            if w:
                w.deleteLater()

    def _btn(self, text: str, handler, r: int, c: int, rs: int = 1, cs: int = 1):
        b = QPushButton(text)
        b.setFocusPolicy(Qt.NoFocus)

        def wrapped(checked=False):
            # 키보드 조작 중 타겟 유지
            if self.target is not None:
                self.target.setFocus(Qt.OtherFocusReason)
            handler()
            self.requestReposition.emit() 

        b.clicked.connect(wrapped)
        self.grid.addWidget(b, r, c, rs, cs)
        return b
    
    def _rebuild(self):
        self._clear_grid()

        if self.mode == "num":
            self.title.setText("숫자 키패드")
            keys = [
                ["7", "8", "9"],
                ["4", "5", "6"],
                ["1", "2", "3"],
                [".", "0", "-"],
            ]
            for r in range(4):
                for c in range(3):
                    k = keys[r][c]
                    self._btn(k, lambda kk=k: self.on_key(kk), r, c)

            self._btn("⌫", self.on_backspace, 0, 3)
            self._btn("Clear", self.on_clear, 1, 3)
            self._btn("Enter", self.on_enter, 2, 3, 2, 1)
            self._btn("닫기", self.on_close, 4, 0, 1, 4)

        else:
            self.title.setText("전체 키보드 (한/영)")
            # 한/영 모드에 따른 키캡
            if self.lang == "EN":
                row1 = list("qwertyuiop")
                row2 = list("asdfghjkl")
                row3 = list("zxcvbnm")
            else:
                # 두벌식 기본 자모(모바일 느낌): 초성/중성 자모를 직접 찍어 조합
                row1 = list("ㅂㅈㄷㄱㅅㅛㅕㅑㅐㅔ")
                row2 = list("ㅁㄴㅇㄹㅎㅗㅓㅏㅣ")
                row3 = list("ㅋㅌㅊㅍㅠㅜㅡ")

            # Row1
            for i, k in enumerate(row1):
                self._btn(k, lambda kk=k: self.on_key(kk), 0, i)

            # Row2
            for i, k in enumerate(row2):
                self._btn(k, lambda kk=k: self.on_key(kk), 1, i)

            # Row3 + 기능키
            self._btn("Shift", self.on_shift, 2, 0, 1, 2)
            col = 2
            for k in row3:
                self._btn(k, lambda kk=k: self.on_key(kk), 2, col)
                col += 1
            self._btn("⌫", self.on_backspace, 2, col, 1, 2)

            # Row4
            self._btn("한/영", self.on_lang, 3, 0, 1, 2)
            self._btn("Space", lambda: self.on_key(" "), 3, 2, 1, 5)
            self._btn("Enter", self.on_enter, 3, 7, 1, 2)
            self._btn("닫기", self.on_close, 3, 9, 1, 1)

            # Row5 (숫자/기호 최소)
            self._btn("1", lambda: self.on_key("1"), 4, 0)
            self._btn("2", lambda: self.on_key("2"), 4, 1)
            self._btn("3", lambda: self.on_key("3"), 4, 2)
            self._btn("4", lambda: self.on_key("4"), 4, 3)
            self._btn("5", lambda: self.on_key("5"), 4, 4)
            self._btn(".", lambda: self.on_key("."), 4, 5)
            self._btn("_", lambda: self.on_key("_"), 4, 6)
            self._btn("-", lambda: self.on_key("-"), 4, 7)
            self._btn("Clear", self.on_clear, 4, 8, 1, 2)

        self.adjustSize()

    def _ensure_target(self) -> bool:
        return self.target is not None

    def _remove_composing_from_target(self):
        """QLineEdit에 덮어썼던 조합 문자열을 삭제한다."""
        if not self._ensure_target():
            return
        if self._composing_len <= 0:
            return
        # 커서 앞에서 composing_len 만큼 삭제
        t = self.target.text()
        cur = self.target.cursorPosition()
        start = max(0, cur - self._composing_len)
        new = t[:start] + t[cur:]
        self.target.setText(new)
        self.target.setCursorPosition(start)
        self._composing_len = 0

    def _insert_text(self, s: str):
        if not self._ensure_target() or s == "":
            return
        t = self.target.text()
        cur = self.target.cursorPosition()
        new = t[:cur] + s + t[cur:]
        self.target.setText(new)
        self.target.setCursorPosition(cur + len(s))

    def _set_composing(self, s: str):
        """조합중 문자를 QLineEdit에 '덮어쓰기'로 표현"""
        self._remove_composing_from_target()
        if s:
            self._insert_text(s)
            self._composing_len = len(s)

    def _commit(self, s: str):
        """조합중 문자는 지우고, 커밋 문자를 넣는다."""
        self._remove_composing_from_target()
        if s:
            self._insert_text(s)

    def on_key(self, k: str):
        if not self._ensure_target():
            return

        # 숫자 모드는 그대로 삽입
        if self.mode == "num":
            self._insert_text(k)
            return

        # full 모드
        if self.lang == "EN":
            ch = k
            if len(ch) == 1 and ch.isalpha():
                if self.shift:
                    ch = ch.upper()
                else:
                    ch = ch.lower()
            # EN 입력 시 한글 조합은 flush하고 그대로 넣기
            if self.composer.has_composing():
                self._commit(self.composer.composing_text())
                self.composer.reset()
            self._insert_text(ch)
            # 일반 키 입력 후 shift는 자동 해제(모바일 스타일)
            if self.shift:
                self.shift = False
                self._rebuild()
            return

        # KO 입력: 자모 조합
        commit, comp = self.composer.input_char(k)
        if commit:
            self._commit(commit)
        self._set_composing(comp)

    def on_backspace(self):
        if not self._ensure_target():
            return

        if self.mode == "num":
            # 커서 앞 1글자 삭제
            t = self.target.text()
            cur = self.target.cursorPosition()
            if cur <= 0:
                return
            new = t[:cur-1] + t[cur:]
            self.target.setText(new)
            self.target.setCursorPosition(cur-1)
            return

        # full: 한글 조합 상태를 먼저 backspace
        remove_real, comp = self.composer.backspace()
        if remove_real:
            # 조합 문자열이 화면에 있으면 제거 후, 실제 문자 1개 삭제
            if self._composing_len > 0:
                self._remove_composing_from_target()
                return
            t = self.target.text()
            cur = self.target.cursorPosition()
            if cur <= 0:
                return
            new = t[:cur-1] + t[cur:]
            self.target.setText(new)
            self.target.setCursorPosition(cur-1)
        else:
            self._set_composing(comp)

    def on_clear(self):
        if not self._ensure_target():
            return
        self.target.clear()
        self.composer.reset()
        self._composing_len = 0

    def on_enter(self):
        if not self._ensure_target():
            return
        # 조합중이면 확정
        if self.composer.has_composing():
            self._commit(self.composer.composing_text())
            self.composer.reset()
            self._composing_len = 0
        # QLineEdit의 returnPressed를 강제로 emit
        try:
            self.target.returnPressed.emit()
        except Exception:
            pass

    def on_close(self):
        self.requestHide.emit()

    def on_shift(self):
        self.shift = not self.shift
        self._rebuild()

    def on_lang(self):
        # 조합중이면 확정하고 모드 전환
        if self.composer.has_composing():
            self._commit(self.composer.composing_text())
            self.composer.reset()
            self._composing_len = 0
        self.lang = "EN" if self.lang == "KO" else "KO"
        self.shift = False
        self._rebuild()


class VirtualKeyboardManager(QObject):
    """
    전역 키보드 1개만 공유.
    - 창마다 Manager는 여러 개여도, 실제 VirtualKeyboard는 1개만 존재.
    - 포커스가 바뀌면 owner만 교체해서 "키보드 2개 뜨는 현상" 제거.
    """
    _GLOBAL_KB: VirtualKeyboard | None = None
    _GLOBAL_OWNER: Union["VirtualKeyboardManager", None] = None

    def __init__(self, parent_window: QWidget):
        super().__init__(parent_window)
        self.parent_window = parent_window

        # ✅ 전역 키보드 1개만 만들기
        if VirtualKeyboardManager._GLOBAL_KB is None:
            VirtualKeyboardManager._GLOBAL_KB = VirtualKeyboard(parent_window)
            # requestHide는 "현재 owner"의 hide를 호출하게 연결 (1회만)
            VirtualKeyboardManager._GLOBAL_KB.requestHide.connect(
                lambda: (VirtualKeyboardManager._GLOBAL_OWNER.hide()
                         if VirtualKeyboardManager._GLOBAL_OWNER else None)
            )

        self.kb = VirtualKeyboardManager._GLOBAL_KB

        self._targets: dict[QLineEdit, str] = {}
        self._hide_timer = QTimer()
        self._hide_timer.setSingleShot(True)
        self._hide_timer.timeout.connect(self.hide)

        self._suppress_until_ms = 0
        self._suppress_target = None

    def register(self, line_edit: QLineEdit, mode: str):
        self._targets[line_edit] = mode
        line_edit.installEventFilter(self)

    def eventFilter(self, obj, event):
        if isinstance(obj, QLineEdit):
            if event.type() == QEvent.FocusIn:
                now_ms = int(time.time() * 1000)

                if (self._suppress_target is obj) and (now_ms < self._suppress_until_ms):
                    return False

                self._suppress_target = None
                self._suppress_until_ms = 0

                mode = self._targets.get(obj, "num")
                self.show_for(obj, mode)
                return False

            if event.type() == QEvent.FocusOut:
                # 터치에서 오판 많음 → 자동 숨김 비활성화
                return False

        return super().eventFilter(obj, event)

    def show_for(self, target: QLineEdit, mode: str):
        self._hide_timer.stop()

        # ✅ 기존 owner가 있으면 그 owner의 hide_timer도 멈추고, 키보드 상태를 우리가 가져온다.
        prev = VirtualKeyboardManager._GLOBAL_OWNER
        if prev is not None and prev is not self:
            prev._hide_timer.stop()

        VirtualKeyboardManager._GLOBAL_OWNER = self

        # ✅ 키보드를 현재 창에 “붙이기” (parent_window가 다르면 재부모화)
        #    (네가 찾은 해결: parent_window 붙여야 좌상단 고정 문제가 줄었지)
        if self.kb.parent() is not self.parent_window:
            self.kb.setParent(self.parent_window)
            self.kb.setWindowFlags(Qt.Tool | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
            self.kb.setAttribute(Qt.WA_ShowWithoutActivating, True)

        self.kb.attach(target, mode)
        self.kb.adjustSize()
        QApplication.processEvents()  # ✅ 네가 찾은 정답 유지

        kb_w = self.kb.width()
        kb_h = self.kb.height()

        below = target.mapToGlobal(QPoint(0, target.height() + 6))
        above = target.mapToGlobal(QPoint(0, -kb_h - 6))

        screen = QGuiApplication.screenAt(below) or QGuiApplication.primaryScreen()
        avail = screen.availableGeometry()

        x = below.x()
        y = below.y()
        if y + kb_h > avail.bottom():
            y = above.y()

        # clamp
        if x + kb_w > avail.right():
            x = avail.right() - kb_w
        if x < avail.left():
            x = avail.left()
        if y + kb_h > avail.bottom():
            y = avail.bottom() - kb_h
        if y < avail.top():
            y = avail.top()

        self.kb.move(x, y)
        if not self.kb.isVisible():
            self.kb.show()
        self.kb.raise_()

    def hide(self):
        # 전역 owner가 나 자신일 때만 실제로 숨김
        if VirtualKeyboardManager._GLOBAL_OWNER is self:
            if self.kb.target is not None:
                self._suppress_target = self.kb.target
                self._suppress_until_ms = int(time.time() * 1000) + 150
            self.kb.hide()


class CameraSelectDialog(QDialog):
    """카메라 목록을 보여주고, 하나 선택해서 OK하는 다이얼로그."""
    def __init__(self, camera_infos, parent=None):
        super().__init__(parent)

        self.setWindowTitle("카메라 연결")
        self.resize(400, 250)
        self.camera_infos = camera_infos
        self.selected_index = None

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(10, 10, 10, 10)
        main_layout.setSpacing(8)

        label = QLabel("확인된 카메라 목록...")
        main_layout.addWidget(label)

        self.list_widget = QListWidget()
        self.list_widget.setSelectionMode(QListWidget.SingleSelection)
        main_layout.addWidget(self.list_widget)

        if not camera_infos:
            self.list_widget.addItem("연결된 카메라가 없습니다.")
            self.list_widget.setEnabled(False)
        else:
            for i, info in enumerate(camera_infos):
                try:
                    text = f"{i}: {info.modelName} | IP: {info.ipAddress} | SN: {info.serialNumber}"
                except Exception:
                    text = f"{i}: {info}"
                self.list_widget.addItem(text)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch(1)

        self.btn_connect = QPushButton("연결")
        self.btn_connect.setFixedWidth(100)
        self.btn_connect.setStyleSheet(
            "QPushButton { background-color: #4CAF50; color:white; padding:6px; }"
        )
        self.btn_connect.clicked.connect(self.on_connect_clicked)

        if not camera_infos:
            self.btn_connect.setEnabled(False)

        btn_layout.addWidget(self.btn_connect)
        main_layout.addLayout(btn_layout)

        self.list_widget.itemDoubleClicked.connect(self.on_connect_clicked)

    def on_connect_clicked(self, *args):
        row = self.list_widget.currentRow()
        if row < 0:
            QMessageBox.warning(self, "선택 필요", "연결할 카메라를 선택하세요.")
            return
        self.selected_index = row
        self.accept()


class CompareSaveDialog(QDialog):
    """비교 데이터 저장(번호/이름 입력 + 목록 표시)."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("비교 데이터 저장")
        self.resize(720, 420)

        os.makedirs(COMPARE_DIR, exist_ok=True)

        main = QVBoxLayout(self)
        main.setContentsMargins(10, 10, 10, 10)
        main.setSpacing(8)

        # 상단: 저장되어 있는 목록 + 번호로 검색
        top_row = QHBoxLayout()
        lbl = QLabel("저장되어 있는 데이터 목록...")
        lbl.setStyleSheet("color:white;")
        top_row.addWidget(lbl)

        top_row.addStretch(1)

        self.edit_search = QLineEdit()
        self.edit_search.setPlaceholderText("번호로 검색")
        self.edit_search.setFixedWidth(220)
        top_row.addWidget(self.edit_search)
        main.addLayout(top_row)

        self.edit_search.setFocusPolicy(Qt.ClickFocus)

        self.list_widget = QListWidget()
        main.addWidget(self.list_widget)

        # 하단 입력
        form = QHBoxLayout()
        lbl_num = QLabel("번호")
        lbl_num.setStyleSheet("color:white;")
        form.addWidget(lbl_num)
        self.edit_number = QLineEdit()
        self.edit_number.setFixedWidth(140)
        form.addWidget(self.edit_number)

        lbl_name = QLabel("이름")
        lbl_name.setStyleSheet("color:white;")
        form.addWidget(lbl_name)
        self.edit_name = QLineEdit()
        self.edit_name.setFixedWidth(260)
        form.addWidget(self.edit_name)

        form.addStretch(1)

        self.btn_save = QPushButton("저장")
        self.btn_save.setFixedSize(140, 48)
        self.btn_save.setStyleSheet("QPushButton { background-color:#4CAF50; color:white; font-size:16px; }")
        form.addWidget(self.btn_save)
        main.addLayout(form)

        # ===== UI 크기/폰트 개선 =====
        self.list_widget.setStyleSheet("font-size:16px;")  # ✅ 목록 글자 키우기

        edit_style = "QLineEdit { font-size:16px; padding:6px; }"
        self.edit_search.setStyleSheet(edit_style)
        self.edit_number.setStyleSheet(edit_style)
        self.edit_name.setStyleSheet(edit_style)

        self.edit_search.setFixedHeight(36)
        self.edit_number.setFixedHeight(40)  # ✅ 입력칸 높이 키우기
        self.edit_name.setFixedHeight(40)

        self.edit_number.setFixedWidth(180)  # ✅ 입력칸 폭 키우기
        self.edit_name.setFixedWidth(360)

        # ✅ 엔터키가 저장(디폴트 버튼)으로 동작하는 것 차단
        self.btn_save.setAutoDefault(False)
        self.btn_save.setDefault(False)

        self.lbl_hint = QLabel("번호와 이름을 입력하면  번호_제품명  폴더에 저장됩니다...")
        self.lbl_hint.setStyleSheet("color:white;")
        main.addWidget(self.lbl_hint)

        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color:red; font-weight:bold;")
        main.addWidget(self.lbl_err)

        # 이벤트
        self.btn_save.clicked.connect(self.on_save_clicked)
        self.edit_search.returnPressed.connect(self.on_search)
        self.edit_number.textChanged.connect(self.on_number_changed)
        self.list_widget.itemClicked.connect(self.on_item_clicked_fill_form)

        self._populate_list()

        self.selected_folder_name: str | None = None
        self.final_folder_name: str | None = None
        self.overwrite_confirmed: bool = False

        # 배경
        pal = self.palette()
        pal.setColor(QPalette.Window, QColor(70, 70, 70))
        self.setAutoFillBackground(True)
        self.setPalette(pal)

        # ============================================================
        # Virtual Keyboard (Dialog)
        # - 번호/검색: num
        # - 이름: full
        # ============================================================
        self.vkb = VirtualKeyboardManager(self)
        self.vkb.register(self.edit_search, "num")
        self.vkb.register(self.edit_number, "num")
        self.vkb.register(self.edit_name, "full")  # ✅ 이름은 전체 키보드

        QTimer.singleShot(0, lambda: self.list_widget.setFocus())


    def _populate_list(self):
        self.list_widget.clear()
        for f in list_compare_folders():
            self.list_widget.addItem(f)

    def on_search(self):
        num = self.edit_search.text().strip()
        f = find_folder_by_number(num)
        if f is None:
            self.lbl_err.setText("해당 번호로 시작하는 폴더를 찾지 못했습니다.")
            return
        self.lbl_err.setText("")
        # 목록에서 선택
        items = self.list_widget.findItems(f, Qt.MatchExactly)
        if items:
            self.list_widget.setCurrentItem(items[0])

    def on_number_changed(self):
        num = self.edit_number.text().strip()
        if num == "":
            self.lbl_hint.setText("번호와 이름을 입력하면  번호_제품명  폴더에 저장됩니다...")
            return
        f = find_folder_by_number(num)
        if f:
            self.lbl_hint.setText(f"⚠ 지정된 번호({num})에 데이터가 존재합니다: {f}")
        else:
            self.lbl_hint.setText("번호와 이름을 입력하면  번호_제품명  폴더에 저장됩니다...")

    def on_save_clicked(self):
        num = sanitize_folder_component(self.edit_number.text())
        name = sanitize_folder_component(self.edit_name.text())

        if num == "" or name == "":
            self.lbl_err.setText("번호 또는 이름이 지정되지 않았습니다.")
            return

        folder_name = f"{num}_{name}"
        folder_path = os.path.join(COMPARE_DIR, folder_name)

        # 중복 번호 처리: 같은 번호_* 폴더가 이미 있으면 경고 + 덮어쓰기 여부
        existing = find_folder_by_number(num)
        if existing is not None and existing != folder_name:
            # 번호는 같지만 이름이 다를 수 있음 → 번호 충돌로 봄
            reply = QMessageBox.question(
                self,
                "경고!",
                "지정된 번호에 데이터가 존재합니다!\n덮어쓰시겠습니까?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                self.lbl_err.setText("저장이 취소되었습니다.")
                return
            # 기존 번호 폴더 삭제 후 새 폴더로 저장
            try:
                shutil.rmtree(os.path.join(COMPARE_DIR, existing))
            except Exception:
                pass

        elif os.path.isdir(folder_path):
            # 정확히 같은 폴더가 존재 → 덮어쓰기 질문
            reply = QMessageBox.question(
                self,
                "경고!",
                "지정된 번호에 데이터가 존재합니다!\n덮어쓰시겠습니까?",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                self.lbl_err.setText("저장이 취소되었습니다.")
                return
            try:
                shutil.rmtree(folder_path)
            except Exception:
                pass

        self.final_folder_name = folder_name
        self.overwrite_confirmed = True
        self.accept()

    def on_item_clicked_fill_form(self, item):
        """저장된 폴더 클릭 시 번호/이름 입력칸 자동 채움."""
        if item is None:
            return
        text = (item.text() or "").strip()
        if text.startswith("("):
            return

        # 폴더명: "번호_이름" 형태 가정
        if "_" in text:
            num, name = text.split("_", 1)
        else:
            num, name = text, ""

        self.edit_number.setText(num)
        self.edit_name.setText(name)
        self.lbl_err.setText("")


class CompareLoadDialog(QDialog):
    """비교 데이터 불러오기(목록 표시 + 번호 검색)."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("비교 데이터 불러오기")
        self.resize(720, 420)

        os.makedirs(COMPARE_DIR, exist_ok=True)

        main = QVBoxLayout(self)
        main.setContentsMargins(10, 10, 10, 10)
        main.setSpacing(8)

        # 상단: 저장되어 있는 목록 + 번호로 검색
        top_row = QHBoxLayout()
        lbl = QLabel("저장되어 있는 데이터 목록...")
        lbl.setStyleSheet("color:white;")
        top_row.addWidget(lbl)

        top_row.addStretch(1)

        self.edit_search = QLineEdit()
        self.edit_search.setPlaceholderText("번호로 검색")
        self.edit_search.setFixedWidth(220)
        top_row.addWidget(self.edit_search)
        main.addLayout(top_row)

        self.edit_search.setFocusPolicy(Qt.ClickFocus)

        self.list_widget = QListWidget()
        main.addWidget(self.list_widget)

        bottom = QHBoxLayout()
        self.lbl_hint = QLabel("데이터를 지정해 주세요.")
        self.lbl_hint.setStyleSheet("color:white;")
        bottom.addWidget(self.lbl_hint)

        bottom.addStretch(1)

        self.btn_load = QPushButton("불러오기")
        self.btn_load.setFixedSize(140, 48)
        self.btn_load.setStyleSheet("QPushButton { background-color:#4CAF50; color:white; font-size:16px; }")
        bottom.addWidget(self.btn_load)

        # ===== UI 크기/폰트 개선 =====
        self.list_widget.setStyleSheet("font-size:16px;")  # ✅ 목록 글자 키우기
        self.edit_search.setStyleSheet("QLineEdit { font-size:16px; padding:6px; }")
        self.edit_search.setFixedHeight(36)

        # ✅ 엔터키가 불러오기(디폴트 버튼)로 동작하는 것 차단
        self.btn_load.setAutoDefault(False)
        self.btn_load.setDefault(False)

        # ✅ 더블클릭으로 불러오기
        self.list_widget.itemDoubleClicked.connect(lambda *_: self.on_load_clicked())

        main.addLayout(bottom)

        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color:red; font-weight:bold;")
        main.addWidget(self.lbl_err)

        self.btn_load.clicked.connect(self.on_load_clicked)
        self.edit_search.returnPressed.connect(self.on_search)

        self._populate_list()

        self.selected_folder: str | None = None

        pal = self.palette()
        pal.setColor(QPalette.Window, QColor(70, 70, 70))
        self.setAutoFillBackground(True)
        self.setPalette(pal)

        # ============================================================
        # Virtual Keyboard (Dialog)
        # - 검색: num
        # ============================================================
        self.vkb = VirtualKeyboardManager(self)
        self.vkb.register(self.edit_search, "num")

        QTimer.singleShot(0, lambda: self.list_widget.setFocus())


    def _populate_list(self):
        self.list_widget.clear()
        folders = list_compare_folders()
        if not folders:
            self.list_widget.addItem("(저장된 비교 데이터가 없습니다)")
            self.list_widget.setEnabled(False)
            self.btn_load.setEnabled(False)
        else:
            for f in folders:
                self.list_widget.addItem(f)

    def on_search(self):
        num = self.edit_search.text().strip()
        f = find_folder_by_number(num)
        if f is None:
            self.lbl_err.setText("해당 번호로 시작하는 폴더를 찾지 못했습니다.")
            return
        self.lbl_err.setText("")
        items = self.list_widget.findItems(f, Qt.MatchExactly)
        if items:
            self.list_widget.setCurrentItem(items[0])

    def on_load_clicked(self):
        if not self.list_widget.isEnabled():
            self.lbl_err.setText("불러올 데이터가 없습니다.")
            return
        row = self.list_widget.currentRow()
        if row < 0:
            self.lbl_err.setText("데이터를 지정해 주세요.")
            return
        folder = self.list_widget.currentItem().text().strip()
        if folder.startswith("("):
            self.lbl_err.setText("데이터를 지정해 주세요.")
            return
        self.selected_folder = folder
        self.accept()


class ExcelTableModel(QAbstractTableModel):
    def __init__(self, headers: list[str], rows: list[list], parent=None):
        super().__init__(parent)
        self.headers = headers or []
        self.rows = rows or []

    def rowCount(self, parent=QModelIndex()):
        return len(self.rows)

    def columnCount(self, parent=QModelIndex()):
        return len(self.headers)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        r = index.row()
        c = index.column()
        v = self.rows[r][c] if (0 <= r < len(self.rows) and 0 <= c < len(self.headers)) else None

        # 표시
        if role == Qt.DisplayRole:
            return "" if v is None else str(v)

        # 정렬(숫자는 숫자로)
        if role == Qt.UserRole:
            if v is None:
                return ""
            try:
                return float(v)
            except Exception:
                return str(v)

        # 정렬/정렬(숫자 우측, 텍스트 좌측, V/X 중앙)
        if role == Qt.TextAlignmentRole:
            header = self.headers[c] if c < len(self.headers) else ""
            text = "" if v is None else str(v)

            if header in ("합격/불합격", "판정 소스(Point)"):
                return Qt.AlignCenter

            # 숫자처럼 보이면 우측 정렬
            try:
                float(text)
                return Qt.AlignVCenter | Qt.AlignRight
            except Exception:
                return Qt.AlignVCenter | Qt.AlignLeft

        # V/X 컬러 강조
        if role == Qt.ForegroundRole:
            header = self.headers[c] if c < len(self.headers) else ""
            if header == "합격/불합격":
                if str(v).strip().upper() == "V":
                    return QColor("green")
                if str(v).strip().upper() == "X":
                    return QColor("red")

        if role == Qt.FontRole:
            header = self.headers[c] if c < len(self.headers) else ""
            if header == "합격/불합격":
                f = Font(bold=True)  # openpyxl Font 말고 Qt Font가 아니라서 아래처럼
                # Qt용 폰트는 PySide6.QtGui.QFont 써야 함
                from PySide6.QtGui import QFont
                qf = QFont()
                qf.setBold(True)
                qf.setPointSize(11)
                return qf

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            if 0 <= section < len(self.headers):
                return self.headers[section]
            return ""
        return str(section + 1)

    def sort(self, column, order=Qt.AscendingOrder):
        if not (0 <= column < len(self.headers)):
            return
        self.layoutAboutToBeChanged.emit()
        reverse = (order == Qt.DescendingOrder)
        self.rows.sort(key=lambda row: ("" if column >= len(row) else row[column]), reverse=reverse)
        self.layoutChanged.emit()


class ExcelLogViewerDialog(QDialog):
    """xlsx 파일을 읽어서 UI 내부(QTableView)로 보여주는 뷰어."""
    def __init__(self, xlsx_path: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Excel Log Viewer - {os.path.basename(xlsx_path)}")
        self.resize(980, 520)

        self.xlsx_path = xlsx_path

        main = QVBoxLayout(self)
        main.setContentsMargins(10, 10, 10, 10)
        main.setSpacing(8)

        top = QHBoxLayout()
        self.lbl_path = QLabel(xlsx_path)
        self.lbl_path.setStyleSheet("color:white;")
        top.addWidget(self.lbl_path)

        top.addStretch(1)

        self.btn_reload = QPushButton("Reload")
        self.btn_reload.setFixedSize(110, 34)
        self.btn_reload.setStyleSheet("QPushButton { background-color:#0e5a7a; color:white; }")
        self.btn_reload.clicked.connect(self.reload)
        top.addWidget(self.btn_reload)

        main.addLayout(top)

        self.table = QTableView()
        self.table.setStyleSheet("background-color:white; color:black;")
        self.table.setAlternatingRowColors(True)

        # ✅ 정렬 허용
        self.table.setSortingEnabled(True)

        # ✅ ResizeToContents는 폭이 튀는 경우가 많아서 Interactive 권장
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)

        # 보기 편한 기본 폭(필요한 열만)
        self.table.horizontalHeader().setMinimumSectionSize(60)

        self.table.verticalHeader().setVisible(False)
        main.addWidget(self.table)

        pal = self.palette()
        pal.setColor(QPalette.Window, QColor(70, 70, 70))
        self.setAutoFillBackground(True)
        self.setPalette(pal)

        self.reload()

    def _read_xlsx(self) -> tuple[list[str], list[list]]:
        if openpyxl is None:
            return (["openpyxl 미설치"], [["pip install openpyxl"]])

        if not os.path.isfile(self.xlsx_path):
            return (["파일 없음"], [[self.xlsx_path]])

        wb = openpyxl.load_workbook(self.xlsx_path, data_only=True, read_only=True)
        ws = wb.active

        # 1행을 헤더로
        headers = []
        for c in range(1, ws.max_column + 1):
            headers.append(ws.cell(1, c).value or "")

        rows = []
        for r in range(2, ws.max_row + 1):
            row = []
            for c in range(1, ws.max_column + 1):
                row.append(ws.cell(r, c).value)
            rows.append(row)

        return headers, rows

    def reload(self):
        headers, rows = self._read_xlsx()
        model = ExcelTableModel(headers, rows, self)
        self.table.setModel(model)


class LogFileLoadDialog(QDialog):
    """error_log 폴더의 날짜별 xlsx 목록 표시 + 검색 + 더블클릭 로드."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Load Other Log")
        self.resize(720, 420)

        os.makedirs(ERROR_LOG_DIR, exist_ok=True)
        self.edit_search.textChanged.connect(self.on_search_live)  # ✅ 추가

        main = QVBoxLayout(self)
        main.setContentsMargins(10, 10, 10, 10)
        main.setSpacing(8)

        top = QHBoxLayout()
        lbl = QLabel("error_log 폴더의 로그 파일(.xlsx)")
        lbl.setStyleSheet("color:white;")
        top.addWidget(lbl)

        top.addStretch(1)

        self.edit_search = QLineEdit()
        self.edit_search.setPlaceholderText("검색 (예: 20251218)")
        self.edit_search.setFixedWidth(260)
        self.edit_search.setStyleSheet("QLineEdit { font-size:16px; padding:6px; }")
        self.edit_search.setFixedHeight(36)
        self.edit_search.setFocusPolicy(Qt.ClickFocus)
        top.addWidget(self.edit_search)

        main.addLayout(top)
        
        self.edit_search.textChanged.connect(self.on_search_live)

        self.list_widget = QListWidget()
        self.list_widget.setStyleSheet("font-size:16px;")
        main.addWidget(self.list_widget)

        bottom = QHBoxLayout()
        self.lbl_hint = QLabel("파일을 선택 후 더블클릭 또는 Load")
        self.lbl_hint.setStyleSheet("color:white;")
        bottom.addWidget(self.lbl_hint)
        bottom.addStretch(1)

        self.btn_load = QPushButton("Load")
        self.btn_load.setFixedSize(140, 48)
        self.btn_load.setStyleSheet("QPushButton { background-color:#4CAF50; color:white; font-size:16px; }")
        self.btn_load.setAutoDefault(False)
        self.btn_load.setDefault(False)
        bottom.addWidget(self.btn_load)

        main.addLayout(bottom)

        self.lbl_err = QLabel("")
        self.lbl_err.setStyleSheet("color:red; font-weight:bold;")
        main.addWidget(self.lbl_err)

        pal = self.palette()
        pal.setColor(QPalette.Window, QColor(70, 70, 70))
        self.setAutoFillBackground(True)
        self.setPalette(pal)

        self.selected_xlsx: str | None = None

        self.btn_load.clicked.connect(self.on_load_clicked)
        self.edit_search.returnPressed.connect(self.on_search)
        self.list_widget.itemDoubleClicked.connect(lambda *_: self.on_load_clicked())

        self._populate()

        # ✅ 검색칸 가상키 (숫자)
        self.vkb = VirtualKeyboardManager(self)
        self.vkb.register(self.edit_search, "num")

        QTimer.singleShot(0, lambda: self.list_widget.setFocus())

    def _populate(self):
        self.list_widget.clear()
        files = []
        for name in os.listdir(ERROR_LOG_DIR):
            if name.lower().endswith(".xlsx"):
                files.append(name)
        files.sort(reverse=True)  # 최신 우선

        if not files:
            self.list_widget.addItem("(로그 파일이 없습니다)")
            self.list_widget.setEnabled(False)
            self.btn_load.setEnabled(False)
        else:
            for f in files:
                self.list_widget.addItem(f)

    def on_search(self):
        key = (self.edit_search.text() or "").strip()
        if key == "":
            self.lbl_err.setText("")
            return

        # contains 검색
        for i in range(self.list_widget.count()):
            item = self.list_widget.item(i)
            if item and key in item.text():
                self.list_widget.setCurrentRow(i)
                self.lbl_err.setText("")
                return

        self.lbl_err.setText("검색 결과 없음")

    def on_load_clicked(self):
        if not self.list_widget.isEnabled():
            self.lbl_err.setText("불러올 로그가 없습니다.")
            return

        item = self.list_widget.currentItem()
        if not item:
            self.lbl_err.setText("파일을 선택하세요.")
            return

        name = item.text().strip()
        if name.startswith("("):
            self.lbl_err.setText("파일을 선택하세요.")
            return

        self.selected_xlsx = os.path.join(ERROR_LOG_DIR, name)
        self.accept()

    def on_search_live(self, text: str):
        key = (text or "").strip()
        if key == "":
            # ✅ 비우면 전체 목록 다시 표시 + 에러 지우기
            self.lbl_err.setText("")
            self._populate()
            return

        # ✅ 현재 목록에서 contains 필터 (전체 재구성)
        self.list_widget.clear()
        files = []
        for name in os.listdir(ERROR_LOG_DIR):
            if name.lower().endswith(".xlsx") and key in name:
                files.append(name)
        files.sort(reverse=True)

        if not files:
            self.list_widget.addItem("(검색 결과 없음)")
            self.list_widget.setEnabled(False)
            self.btn_load.setEnabled(False)
            self.lbl_err.setText("검색 결과 없음")
        else:
            self.list_widget.setEnabled(True)
            self.btn_load.setEnabled(True)
            for f in files:
                self.list_widget.addItem(f)
            self.lbl_err.setText("")
            self.list_widget.setCurrentRow(0)


class ScanViewer(QWidget):
    """
    3D / 2D 스캔 데이터 출력 영역.
    - Point: pyqtgraph GLViewWidget
    - QImage: 내부에 원본 QImage를 저장해두고 resize 시 자동 재스케일
    """
    def __init__(self, parent=None):
        super().__init__(parent)

        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(0, 0, 0, 0)
        self._layout.setSpacing(0)

        self.label = QLabel("스캔 데이터 출력")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("color: white; font-size: 16px;")
        self.label.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._layout.addWidget(self.label)

        self.gl_view = None
        self.gl_scatter = None

        self._current_qimg: QImage | None = None
        self._empty_text = "스캔 데이터 출력"

        palette = self.palette()
        palette.setColor(QPalette.Window, QColor(0, 0, 0))
        self.setAutoFillBackground(True)
        self.setPalette(palette)

    def ensure_gl_view(self):
        if GLViewWidget is None or GLScatterPlotItem is None:
            return False
        if self.gl_view is None:
            self.gl_view = GLViewWidget()
            self.gl_view.setBackgroundColor('k')
            self.gl_view.opts['fov'] = 60
            self.gl_view.opts['elevation'] = 25
            self.gl_view.opts['azimuth'] = 45
            self.gl_view.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self._layout.addWidget(self.gl_view)

        self._current_qimg = None
        self.label.hide()
        self.gl_view.show()
        return True

    def show_pointcloud(self, points: np.ndarray, colors: np.ndarray | None = None):
        if not self.ensure_gl_view():
            self.label.setText("pyqtgraph / PyOpenGL 설치 필요: pip install pyqtgraph PyOpenGL")
            self.label.show()
            if self.gl_view is not None:
                self.gl_view.hide()
            return

        if points is None or points.size == 0:
            self.label.setText("표시할 포인트가 없습니다.")
            self.label.show()
            if self.gl_view is not None:
                self.gl_view.hide()
            return

        pts = points.astype(np.float32)

        max_points = 150_000
        if pts.shape[0] > max_points:
            idx = np.random.choice(pts.shape[0], max_points, replace=False)
            pts = pts[idx]
            if colors is not None:
                colors = colors[idx]

        center = pts.mean(axis=0)
        pts_centered = pts - center

        if colors is not None:
            cols = colors.astype(np.float32)
            if cols.max() > 1.0:
                cols = cols / 255.0
        else:
            cols = np.ones((pts.shape[0], 3), dtype=np.float32)

        if cols.ndim == 2 and cols.shape[1] == 3:
            alpha = np.ones((cols.shape[0], 1), dtype=np.float32)
            cols_rgba = np.concatenate([cols, alpha], axis=1)
        else:
            cols_rgba = cols

        radius = float(np.linalg.norm(pts_centered, axis=1).max())
        if radius > 0:
            self.gl_view.opts['distance'] = radius * 2.5

        point_size = 2.0

        if self.gl_scatter is None:
            self.gl_scatter = GLScatterPlotItem(
                pos=pts_centered,
                color=cols_rgba,
                size=point_size,
                pxMode=True,
                glOptions='opaque',
            )
            self.gl_view.addItem(self.gl_scatter)
        else:
            self.gl_scatter.setData(pos=pts_centered, color=cols_rgba, size=point_size, pxMode=True)
            self.gl_scatter.setGLOptions('opaque')

        if pg is not None:
            self.gl_view.opts['center'] = pg.Vector(0, 0, 0)

        self.gl_view.show()

    def show_qimage(self, qimg: QImage | None, empty_text: str = "스캔 데이터 출력"):
        self._empty_text = empty_text
        self._current_qimg = qimg

        if self.gl_view is not None:
            self.gl_view.hide()

        self._render_qimage()

    def _render_qimage(self):
        if self._current_qimg is None:
            self.label.setPixmap(QPixmap())
            self.label.setText(self._empty_text)
            self.label.setAlignment(Qt.AlignCenter)
            self.label.show()
            return

        pixmap = QPixmap.fromImage(self._current_qimg)
        scaled = pixmap.scaled(self.label.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.label.setPixmap(scaled)
        self.label.setText("")
        self.label.setAlignment(Qt.AlignCenter)
        self.label.show()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        # ✅ 이미지 모드일 때만 재스케일
        if self._current_qimg is not None and self.label.isVisible():
            self._render_qimage()


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("3D 카메라 스캔 인터페이스 (Improved UI)")
        self.resize(1600, 900)          # ✅ 초기 창 크게
        self.setMinimumSize(1500, 850)

        os.makedirs(COMPARE_DIR, exist_ok=True)

        # -------------------------
        # 카메라 & 데이터 상태(기존 유지)
        # -------------------------
        self.camera = None
        self.camera_connected = False

        self.last_pointcloud = None
        self.last_colors = None
        self.last_depth_qimage: QImage | None = None
        self.last_depth_array: np.ndarray | None = None
        self.last_color_qimage: QImage | None = None

        self.compare_pointcloud = None
        self.compare_colors = None
        self.compare_depth_qimage: QImage | None = None
        self.compare_depth_array: np.ndarray | None = None
        self.compare_color_qimage: QImage | None = None

        self.show_diff_overlay = False
        self.last_capture_dt: datetime | None = None
        self.last_capture_name: str | None = None
        self.current_compare_name: str | None = None

        # -------------------------
        # Theme
        # -------------------------
        pal = self.palette()
        pal.setColor(QPalette.Window, QColor(45, 45, 45))
        self.setPalette(pal)

        self.setStyleSheet("""
            QMainWindow { background:#2d2d2d; }
            QGroupBox {
                color:white; border:1px solid #555; border-radius:8px;
                margin-top:10px; padding:10px;
            }
            QGroupBox::title { subcontrol-origin: margin; left:10px; padding:0 6px; }
            QLabel { color:white; }
            QLineEdit { background:white; color:black; padding:6px; border-radius:6px; }
            QTextEdit { background:white; color:black; border-radius:8px; }
            QPushButton {
                background:#3f4c55; color:white; padding:8px 10px; border-radius:8px;
                font-size:13px;
            }
            QPushButton:checked { background:#0e5a7a; }
            QPushButton:disabled { background:#4a6c7a; }
        """)

        # -------------------------
        # UI Build
        # -------------------------
        self._build_toolbar()
        self._build_central_views()
        self._build_settings_dock()
        self._build_log_dock()
        self._build_statusbar()

        # -------------------------
        # Virtual Keyboard 적용(기존 유지)
        # -------------------------
        self.vkb = VirtualKeyboardManager(self)
        self.vkb.register(self.edit_tol_distance, "num")
        self.vkb.register(self.edit_tol_ratio, "num")
        self.vkb.register(self.edit_gain_db, "num")
        self.vkb.register(self.edit_exposure_ms, "num")

        QTimer.singleShot(0, self._apply_fixed_layout_sizes)

        self._fixed_center_h = None

        self._dock_resize_timer = QTimer(self)
        self._dock_resize_timer.setSingleShot(True)
        self._dock_resize_timer.timeout.connect(self._sync_docks_on_resize)

        # 최대화/전체화면일 때 Log가 과하게 커지지 않도록 제한값
        self._min_bottom_h = 160
        self._max_bottom_h = 340
        self._maximized_bottom_ratio = 0.28  # 전체 높이의 28% 정도만 Log에 배정 (취향대로 0.22~0.35 추천)

        if pg is None or GLViewWidget is None:
            self.append_log("[ERROR] pyqtgraph / PyOpenGL 미설치. 3D 뷰어를 사용하려면 pip install pyqtgraph PyOpenGL")

    def showEvent(self, event):
        super().showEvent(event)
        if getattr(self, "_central_sizes_applied", False):
            return
        self._central_sizes_applied = True
        QTimer.singleShot(0, self._force_central_equal)
        QTimer.singleShot(0, self._force_central_equal)  # ✅ 2번 걸어주면 더 안정적

    def _force_central_equal(self):
        sp = getattr(self, "central_splitter", None)
        if sp is None:
            return
        w = sp.width()
        if w <= 10:
            return
        half = w // 2
        sp.setSizes([half, w - half])

    def _build_toolbar(self):
        tb = QToolBar("Main")
        tb.setMovable(False)
        tb.setIconSize(QSize(16, 16))
        tb.setStyleSheet("QToolBar { background:#232323; spacing:6px; border:none; }")
        self.addToolBar(Qt.TopToolBarArea, tb)

        # --- View Mode (기존 btn_* 이름 유지) ---
        self.btn_point = QPushButton("Point")
        self.btn_depth = QPushButton("Depths")
        self.btn_image = QPushButton("Image")
        for b in (self.btn_point, self.btn_depth, self.btn_image):
            b.setCheckable(True)
            b.clicked.connect(self.on_mode_changed)

        self.btn_point.setChecked(True)
        group = QButtonGroup(self)
        group.setExclusive(True)
        group.addButton(self.btn_point)
        group.addButton(self.btn_depth)
        group.addButton(self.btn_image)

        tb.addWidget(self.btn_point)
        tb.addWidget(self.btn_depth)
        tb.addWidget(self.btn_image)
        tb.addSeparator()

        # --- Core Flow (버튼은 생성만 하고, 배치는 다른 패널로 이동) ---
        self.btn_connect = QPushButton("Connect")
        self.btn_connect.clicked.connect(self.connect_camera)

        self.btn_scan = QPushButton("Scan")
        self.btn_scan.clicked.connect(self.scan_pointcloud)
        self.btn_scan.setEnabled(False)

        self.btn_save_files = QPushButton("Save Recent")
        self.btn_save_files.setEnabled(False)
        self.btn_save_files.clicked.connect(self.on_save_files_clicked)

        self.btn_define_compare = QPushButton("Define Compare")
        self.btn_define_compare.clicked.connect(self.on_define_compare_clicked)

        self.btn_save_compare_data = QPushButton("Save Compare")
        self.btn_save_compare_data.clicked.connect(self.on_save_compare_data_clicked)

        self.btn_load_compare_data = QPushButton("Load Compare")
        self.btn_load_compare_data.clicked.connect(self.on_load_compare_data_clicked)

        self.btn_reset_compare_data = QPushButton("Reset Compare")
        self.btn_reset_compare_data.clicked.connect(self.on_reset_compare_data_clicked)

        # --- Logs / Git ---
        self.btn_watch_today_log = QPushButton("Today Log")
        self.btn_watch_today_log.clicked.connect(self.on_watch_today_log_clicked)

        self.btn_load_other_log = QPushButton("Other Log")
        self.btn_load_other_log.clicked.connect(self.on_load_other_log_clicked)

        self.btn_upload_git = QPushButton("Upload Git")
        self.btn_upload_git.clicked.connect(self.on_upload_git_clicked)

        # --- Status labels (이 라벨도 각 패널 헤더로 배치) ---
        self.lbl_toolbar_compare = QLabel("Compare: -")
        self.lbl_toolbar_compare.setStyleSheet("color:#cfcfcf; padding:0 8px;")

        self.label_last_capture = QLabel("Last: -")
        self.label_last_capture.setStyleSheet("color:#cfcfcf; padding-right:8px;")
        # (툴바에는 View Mode만 남김)

    def _build_central_views(self):
        splitter = FixedSplitter(Qt.Horizontal)   # ✅ 중앙 드래그 불가(완전 고정)
        self.central_splitter = splitter          # ✅ 참조 저장
        splitter.setChildrenCollapsible(False)
        splitter.setHandleWidth(1)
        splitter.setChildrenCollapsible(False)
        gb_left = QGroupBox("저장된 비교 대상")
        left_layout = QVBoxLayout(gb_left)

        # ✅ Compare 패널 헤더: Compare 라벨 + 비교데이터 버튼들
        left_header = QHBoxLayout()
        left_header.setContentsMargins(0, 0, 0, 0)
        left_header.setSpacing(6)
        left_header.addWidget(self.lbl_toolbar_compare, 1)
        left_header.addWidget(self.btn_define_compare, 0)
        left_header.addWidget(self.btn_save_compare_data, 0)
        left_header.addWidget(self.btn_load_compare_data, 0)
        left_header.addWidget(self.btn_reset_compare_data, 0)
        left_layout.addLayout(left_header)

        self.compare_viewer = ScanViewer()
        left_layout.addWidget(self.compare_viewer, 1)

        gb_right = QGroupBox("가장 최근에 촬영된 이미지")
        right_layout = QVBoxLayout(gb_right)

        # ✅ Recent 패널 헤더: Last 라벨 + Save Recent / Upload Git
        right_header = QHBoxLayout()
        right_header.setContentsMargins(0, 0, 0, 0)
        right_header.setSpacing(6)
        right_header.addWidget(self.label_last_capture, 1)
        right_header.addWidget(self.btn_save_files, 0)
        right_header.addWidget(self.btn_upload_git, 0)
        right_layout.addLayout(right_header)

        self.scan_viewer = ScanViewer()
        right_layout.addWidget(self.scan_viewer, 1)

        splitter.addWidget(gb_left)
        splitter.addWidget(gb_right)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 1)

        self.setCentralWidget(splitter)

    def _build_settings_dock(self):
        dock = QDockWidget("Settings", self)
        self.dock_settings = dock
        dock.setFeatures(QDockWidget.NoDockWidgetFeatures)
        dock.setAllowedAreas(Qt.LeftDockWidgetArea | Qt.RightDockWidgetArea)
        dock.setFeatures(QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetFloatable)
        self.addDockWidget(Qt.RightDockWidgetArea, dock)

        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(10, 10, 10, 10)
        v.setSpacing(10)

        # ---- Verdict / Diff ----
        gb_verdict = QGroupBox("판정 / Diff")
        vv = QVBoxLayout(gb_verdict)

        self.label_tol_indicator = QLabel("-")
        self.label_tol_indicator.setAlignment(Qt.AlignCenter)
        self.label_tol_indicator.setStyleSheet("background:white; color:black; font-size:34px; font-weight:bold; border-radius:10px;")
        self.label_tol_indicator.setMinimumHeight(70)
        vv.addWidget(self.label_tol_indicator)

        self.btn_diff_toggle = QPushButton("Diff 표시 OFF")
        self.btn_diff_toggle.setCheckable(True)
        self.btn_diff_toggle.clicked.connect(self.on_diff_toggle_clicked)
        vv.addWidget(self.btn_diff_toggle)

        v.addWidget(gb_verdict)

        # ---- Tolerance ----
        gb_tol = QGroupBox("허용 오차")
        form = QFormLayout(gb_tol)
        form.setLabelAlignment(Qt.AlignLeft)
        form.setFormAlignment(Qt.AlignTop)

        self.edit_tol_distance = QLineEdit("10")
        self.edit_tol_ratio = QLineEdit("2")
        form.addRow("오차 거리(mm)", self.edit_tol_distance)
        form.addRow("오차 비율(%)", self.edit_tol_ratio)

        self.btn_apply_tol = QPushButton("Apply Tolerance")
        self.btn_apply_tol.clicked.connect(self.on_apply_tolerance_clicked)
        form.addRow(self.btn_apply_tol)

        v.addWidget(gb_tol)

        # ---- Camera ----
        gb_cam = QGroupBox("카메라 파라미터 (2D/3D 동일)")
        form2 = QFormLayout(gb_cam)

        self.edit_gain_db = QLineEdit("12")
        self.edit_exposure_ms = QLineEdit("6.0")

        form2.addRow("게인(dB)", self.edit_gain_db)
        form2.addRow("노출(ms)", self.edit_exposure_ms)

        self.btn_apply_cam = QPushButton("Apply Camera")
        self.btn_apply_cam.clicked.connect(self.on_apply_camera_params_clicked)
        form2.addRow(self.btn_apply_cam)

        v.addWidget(gb_cam)

        # ---- Quick Clear Log ----
        self.btn_refresh_top = QPushButton("로그 초기화")
        self.btn_refresh_top.clicked.connect(self.on_refresh_clicked)
        v.addWidget(self.btn_refresh_top)

        v.addStretch(1)
        dock.setWidget(w)

    def _build_log_dock(self):
        dock = QDockWidget("Log", self)
        self.dock_log = dock
        dock.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea)
        dock.setFeatures(QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetFloatable)
        self.addDockWidget(Qt.BottomDockWidgetArea, dock)

        w = QWidget()
        v = QVBoxLayout(w)
        v.setContentsMargins(10, 10, 10, 10)
        v.setSpacing(8)

        top = QHBoxLayout()
        btn_clear = QPushButton("Clear")
        btn_clear.clicked.connect(self.on_refresh_clicked)
        top.addWidget(btn_clear)

        btn_copy = QPushButton("Copy")
        def _copy():
            QGuiApplication.clipboard().setText(self.log_text.toPlainText())
        btn_copy.clicked.connect(_copy)
        top.addWidget(btn_copy)

        btn_save = QPushButton("Save .txt")
        def _save():
            path, _ = QFileDialog.getSaveFileName(self, "Save Log", BASE_DIR, "Text Files (*.txt)")
            if path:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(self.log_text.toPlainText())
                self.append_log(f"[INFO] 로그 저장: {path}")
        btn_save.clicked.connect(_save)
        top.addWidget(btn_save)

        # ✅ 로그 엑셀 버튼은 로그 창 상단에 붙임
        top.addSpacing(12)
        top.addWidget(self.btn_watch_today_log)
        top.addWidget(self.btn_load_other_log)

        top.addStretch(1)
        v.addLayout(top)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        v.addWidget(self.log_text, 1)

        dock.setWidget(w)

        # ============================================================
        # ✅ Camera Dock (Connect/Scan) 추가 → Log Dock과 가로로 split
        # ============================================================
        dock_cam = QDockWidget("Camera", self)
        self.dock_camera = dock_cam
        dock_cam.setAllowedAreas(Qt.BottomDockWidgetArea | Qt.TopDockWidgetArea)
        dock_cam.setFeatures(QDockWidget.DockWidgetMovable | QDockWidget.DockWidgetFloatable)
        self.addDockWidget(Qt.BottomDockWidgetArea, dock_cam)
        self.splitDockWidget(dock, dock_cam, Qt.Horizontal)  # ✅ 로그 폭을 반으로

        cw = QWidget()
        cv = QVBoxLayout(cw)
        cv.setContentsMargins(10, 10, 10, 10)
        cv.setSpacing(10)

        # 버튼 크게/명확하게
        self.btn_connect.setMinimumHeight(44)
        self.btn_scan.setMinimumHeight(44)
        cv.addWidget(self.btn_connect)
        cv.addWidget(self.btn_scan)
        cv.addStretch(1)
        dock_cam.setWidget(cw)

    def _build_statusbar(self):
        self.statusBar().setStyleSheet("color:#d0d0d0;")
        self.statusBar().showMessage("Ready")

    def _apply_fixed_layout_sizes(self):
        """
        도킹/스플리터는 show 이후에 Qt가 재배치하면서 크기가 튈 수 있어서
        show 직후(0ms)에 한 번 더 '원하는 자연스러운 기본값'으로 고정한다.
        """

        # --- 우측 Settings dock 폭 고정 ---
        if hasattr(self, "dock_settings") and self.dock_settings is not None:
            # 폭을 고정(드래그로 틀어지는 것 방지)
            self.dock_settings.setMinimumWidth(340)
            self.dock_settings.setMaximumWidth(340)
            self.resizeDocks([self.dock_settings], [340], Qt.Horizontal)

        # --- 하단 Log / Camera 도크 영역 높이 기본값 ---
        if hasattr(self, "dock_log") and self.dock_log is not None:
            # 하단 높이(너무 커지거나 작아지는 것 방지용 기본값)
            self.resizeDocks([self.dock_log], [260], Qt.Vertical)

        # --- 하단 Camera dock 폭 고정 + 로그가 나머지 차지 ---
        if hasattr(self, "dock_camera") and self.dock_camera is not None and hasattr(self, "dock_log"):
            cam_w = 320
            self.dock_camera.setMinimumWidth(cam_w)
            self.dock_camera.setMaximumWidth(cam_w)

            total_w = max(1, self.width())
            log_w = max(200, total_w - cam_w - 40)  # 약간의 여유
            self.resizeDocks([self.dock_log, self.dock_camera], [log_w, cam_w], Qt.Horizontal)

        # --- 중앙 좌/우 기본 반반 (드래그 막았든 말았든 초기값 고정) ---
        if hasattr(self, "central_splitter") and self.central_splitter is not None:
            w = max(1, self.central_splitter.width())
            self.central_splitter.setSizes([w // 2, w // 2])

        QTimer.singleShot(0, self._capture_fixed_center_height)

    def _capture_fixed_center_height(self):
        # 중앙 패널(좌/우 3D 출력 영역)의 "현재 높이"를 고정 기준으로 저장
        if self._fixed_center_h is not None:
            return
        if not hasattr(self, "central_splitter") or self.central_splitter is None:
            return

        h = self.central_splitter.height()
        if h > 50:
            self._fixed_center_h = h
            # 아래 최소 높이도 같이 잡아두면 너무 눌리는 걸 방지할 수 있음(선택)
            # self.central_splitter.setMinimumHeight(300)

    def _is_max_like(self) -> bool:
        return self.isMaximized() or self.isFullScreen()

    def _bottom_docks(self):
        docks = []
        for dw in (getattr(self, "dock_log", None), getattr(self, "dock_camera", None)):
            if dw and self.dockWidgetArea(dw) == Qt.BottomDockWidgetArea:
                docks.append(dw)
        return docks

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if self._fixed_center_h is not None:
            self._dock_resize_timer.start(0)

    def changeEvent(self, event):
        super().changeEvent(event)
        if event.type() == QEvent.WindowStateChange:
            # 최대화/전체화면 진입/해제 시 레이아웃 재정렬
            QTimer.singleShot(0, self._on_window_state_changed)

    def _on_window_state_changed(self):
        # 최대화에서 복원되면, 복원된 상태의 중앙 높이를 다시 "고정 기준"으로 잡아줌
        if not self._is_max_like() and hasattr(self, "central_splitter") and self.central_splitter:
            self._fixed_center_h = self.central_splitter.height()
        self._sync_docks_on_resize()

    def _sync_docks_on_resize(self):
        if self._fixed_center_h is None:
            return
        if not hasattr(self, "central_splitter") or self.central_splitter is None:
            return

        bottom_docks = self._bottom_docks()
        if not bottom_docks:
            return

        # ✅ 최대화/전체화면: 중앙도 자연스럽게 커지게 두고(Log만 적당히)
        if self._is_max_like():
            desired = int(self.height() * self._maximized_bottom_ratio)
            desired = max(self._min_bottom_h, min(self._max_bottom_h, desired))
            self.resizeDocks(bottom_docks, [desired] * len(bottom_docks), Qt.Vertical)
            return

        # ✅ 일반 모드: 중앙 높이 고정 + 늘어난 만큼 Log가 먹게
        center_now = self.central_splitter.height()
        delta = center_now - self._fixed_center_h
        if abs(delta) < 2:
            return

        bottom_now = bottom_docks[0].height()
        new_bottom = max(self._min_bottom_h, bottom_now + delta)

        self.resizeDocks(bottom_docks, [new_bottom] * len(bottom_docks), Qt.Vertical)

        # 레이아웃이 한 템포 늦게 반영되면 한번 더
        if abs(self.central_splitter.height() - self._fixed_center_h) > 2:
            self._dock_resize_timer.start(0)

    # ✅ 툴바/상태 표시 업데이트를 한 곳으로
    def _sync_top_labels(self):
        cmp = self.current_compare_name or "-"
        self.lbl_toolbar_compare.setText(f"Compare: {cmp}")

        if self.last_capture_dt is None:
            self.label_last_capture.setText("Last: -")
        else:
            self.label_last_capture.setText("Last: " + self.last_capture_dt.strftime("%Y-%m-%d %H:%M:%S"))

    # ===== 로그 도우미 =====
    def append_log(self, text: str):
        now = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{now}] {text}")

    # ===== 모드 전환 =====
    def on_mode_changed(self):
        if self.btn_point.isChecked():
            if self.compare_pointcloud is not None:
                self.update_compare_pointcloud_view()

            if self.last_pointcloud is not None:
                if self.show_diff_overlay and self.compare_pointcloud is not None:
                    self.update_recent_pointcloud_diff_view()
                else:
                    self.update_recent_pointcloud_view()

        elif self.btn_depth.isChecked():
            if self.compare_depth_qimage is not None:
                self.update_compare_depth_view()
            if self.last_depth_qimage is not None:
                self.update_recent_depth_view()

        elif self.btn_image.isChecked():
            if self.compare_color_qimage is not None:
                self.update_compare_image_view()
            if self.last_color_qimage is not None:
                self.update_recent_image_view()

    # ===== 로그 초기화 =====
    def on_refresh_clicked(self):
        self.log_text.clear()

    def on_upload_git_clicked(self):
        script_path = os.path.join(BASE_DIR, "UploadGit.py")

        if not os.path.isfile(script_path):
            self.append_log(f"[ERROR] UploadGit.py 를 찾을 수 없습니다: {script_path}")
            return

        self.append_log("[GIT] UploadGit.py 실행...")

        try:
            env = os.environ.copy()
            env["PYTHONUTF8"] = "1"
            env["PYTHONIOENCODING"] = "utf-8"

            result = subprocess.run(
                [sys.executable, "-X", "utf8", script_path],
                cwd=BASE_DIR,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                env=env,
            )

            out = (result.stdout or "").strip()
            err = (result.stderr or "").strip()

            if out:
                for line in out.splitlines():
                    self.append_log(f"[UploadGit.py] {line}")

            if err:
                for line in err.splitlines():
                    self.append_log(f"[UploadGit.py][ERR] {line}")

            if result.returncode == 0:
                self.append_log("[GIT] UploadGit.py 완료 (returncode=0)")
            else:
                self.append_log(f"[GIT] UploadGit.py 실패 (returncode={result.returncode})")

        except Exception as e:
            self.append_log(f"[ERROR] UploadGit.py 실행 중 예외: {e}")

    # ===== 카메라 연결 =====
    def connect_camera(self):
        if not USE_REAL_CAMERA:
            self.append_log("[INFO] (더미) 카메라 연결을 생략합니다. 샘플 파일 모드로 동작합니다.")
            self.camera_connected = True
            self.btn_scan.setEnabled(True)
            return

        try:
            camera_infos = Camera.discover_cameras()

            dlg = CameraSelectDialog(camera_infos, self)
            result = dlg.exec()

            if not camera_infos:
                self.append_log("[ERROR] 검색된 카메라가 없습니다.")
                return

            if result != QDialog.Accepted or dlg.selected_index is None:
                return

            idx = dlg.selected_index

            self.camera = Camera()
            error_status = self.camera.connect(camera_infos[idx])

            if not error_status.is_ok():
                try:
                    show_error(error_status)
                except Exception:
                    pass
                self.append_log(f"[ERROR] 카메라 연결 실패: {error_status}")
                self.camera_connected = False
                self.btn_scan.setEnabled(False)
                return

            self.camera_connected = True
            self.btn_scan.setEnabled(True)

        except Exception as e:
            self.append_log(f"[ERROR] 카메라 연결 중 예외: {e}")
            self.camera_connected = False
            self.btn_scan.setEnabled(False)

    # ===== 스캔 =====
    def scan_pointcloud(self):
        if USE_REAL_CAMERA and not self.camera_connected:
            self.append_log("[ERROR] 카메라가 연결되지 않았습니다.")
            return

        self.append_log("촬영 실행: 2D + Depth + Point 데이터를 스캔합니다.")
        try:
            self.capture_both_depth_and_point()
        except Exception as e:
            self.append_log(f"[ERROR] 캡쳐 중 오류: {e}")
            return

        now_dt = datetime.now()
        self.last_capture_dt = now_dt
        self.last_capture_name = now_dt.strftime("%Y%m%d_%H%M%S")
        self.label_last_capture.setText(now_dt.strftime("%Y-%m-%d %H:%M:%S"))
        self.btn_save_files.setEnabled(True)

        if self.btn_point.isChecked():
            if self.last_pointcloud is None:
                self.append_log("[ERROR] 포인트 데이터가 없습니다.")
                return

            if self.show_diff_overlay and self.compare_pointcloud is not None:
                self.update_recent_pointcloud_diff_view()
            else:
                self.update_recent_pointcloud_view()

        elif self.btn_depth.isChecked():
            if self.last_depth_qimage is None:
                self.append_log("[ERROR] 뎁스 데이터가 없습니다.")
                return
            self.update_recent_depth_view()

        elif self.btn_image.isChecked():
            if self.last_color_qimage is None:
                self.append_log("[ERROR] 2D 이미지가 없습니다.")
                return
            self.update_recent_image_view()

        self.update_tolerance_display()
        self._handle_error_on_scan_if_needed()
        self._sync_top_labels()

    # ===== 최근 스캔 데이터 파일로 저장 =====
    def on_save_files_clicked(self):
        """
        result/날짜시간/ 폴더에
        - PNG : 2D 카메라 이미지
        - TIFF: 원본 뎁스 배열
        - PLY : 포인트 클라우드
        를 저장.
        """
        if self.last_capture_dt is None or self.last_capture_name is None:
            self.append_log("[ERROR] 저장할 스캔이 없습니다. 먼저 Scan을 실행하세요.")
            return

        if (
            self.last_pointcloud is None
            and self.last_depth_array is None
            and self.last_color_qimage is None
        ):
            self.append_log("[ERROR] 저장할 데이터가 없습니다.")
            return

        os.makedirs(RESULT_DIR, exist_ok=True)
        save_dir = os.path.join(RESULT_DIR, self.last_capture_name)
        os.makedirs(save_dir, exist_ok=True)

        base = self.last_capture_name

        # PNG
        if self.last_color_qimage is not None:
            png_path = os.path.join(save_dir, f"{base}.png")
            ok = self.last_color_qimage.save(png_path, "PNG")
            if ok:
                self.append_log(f"PNG(2D 이미지) 저장 완료: {png_path}")
            else:
                self.append_log("[ERROR] PNG 저장 실패")
        else:
            self.append_log("[WARN] PNG 저장 생략: 2D 이미지 없음")

        # TIFF
        if self.last_depth_array is not None:
            tiff_path = os.path.join(save_dir, f"{base}.tiff")
            try:
                saved = False
                depth = self.last_depth_array.astype(np.float32)
                if Image is not None:
                    img = Image.fromarray(depth)
                    img.save(tiff_path, format="TIFF")
                    saved = True
                elif cv2 is not None:
                    cv2.imwrite(tiff_path, depth)
                    saved = True

                if saved:
                    self.append_log(f"TIFF 저장 완료: {tiff_path}")
                else:
                    self.append_log("[ERROR] TIFF 저장 실패: PIL / OpenCV 미설치")
            except Exception as e:
                self.append_log(f"[ERROR] TIFF 저장 중 예외: {e}")
        else:
            self.append_log("[WARN] TIFF 저장 생략: 뎁스 배열 없음")

        # PLY
        if self.last_pointcloud is not None and self.last_pointcloud.size > 0:
            ply_path = os.path.join(save_dir, f"{base}.ply")
            try:
                pts = self.last_pointcloud
                cols = self.last_colors

                if o3d is not None:
                    pcd = o3d.geometry.PointCloud()
                    pcd.points = o3d.utility.Vector3dVector(pts)
                    if cols is not None:
                        cols_to_use = cols.astype(np.float32)
                        if cols_to_use.max() > 1.0:
                            cols_to_use = cols_to_use / 255.0
                        pcd.colors = o3d.utility.Vector3dVector(cols_to_use)
                    o3d.io.write_point_cloud(ply_path, pcd)
                else:
                    n = pts.shape[0]
                    has_color = cols is not None and cols.shape[0] == n

                    with open(ply_path, "w", encoding="utf-8") as f:
                        f.write("ply\n")
                        f.write("format ascii 1.0\n")
                        f.write(f"element vertex {n}\n")
                        f.write("property float x\n")
                        f.write("property float y\n")
                        f.write("property float z\n")
                        if has_color:
                            f.write("property uchar red\n")
                            f.write("property uchar green\n")
                            f.write("property uchar blue\n")
                        f.write("end_header\n")

                        for i in range(n):
                            x, y, z = pts[i]
                            if has_color:
                                c = cols[i].astype(np.float32)
                                if c.max() <= 1.0:
                                    c = c * 255.0
                                r, g, b = [int(np.clip(v, 0, 255)) for v in c]
                                f.write(f"{x} {y} {z} {r} {g} {b}\n")
                            else:
                                f.write(f"{x} {y} {z}\n")

                self.append_log(f"PLY 저장 완료: {ply_path}")
            except Exception as e:
                self.append_log(f"[ERROR] PLY 저장 중 예외: {e}")
        else:
            self.append_log("[WARN] PLY 저장 생략: 포인트 클라우드 없음")

        self.append_log(f"스캔 데이터 저장 완료: {save_dir}")

    # ===== 기존: 비교데이터 "정의" (last_* → compare_*) =====
    def on_define_compare_clicked(self):
        if (
            self.last_pointcloud is None
            and self.last_depth_array is None
            and self.last_color_qimage is None
        ):
            self.append_log("[ERROR] 비교 데이터로 저장할 스캔이 없습니다.")
            return

        self.compare_pointcloud = None if self.last_pointcloud is None else self.last_pointcloud.copy()
        self.compare_colors = None if self.last_colors is None else self.last_colors.copy()

        self.compare_depth_array = None if self.last_depth_array is None else self.last_depth_array.copy()
        self.compare_depth_qimage = None if self.last_depth_qimage is None else self.last_depth_qimage.copy()

        self.compare_color_qimage = None if self.last_color_qimage is None else self.last_color_qimage.copy()

        if self.btn_point.isChecked():
            self.update_compare_pointcloud_view()
        elif self.btn_depth.isChecked():
            self.update_compare_depth_view()
        elif self.btn_image.isChecked():
            self.update_compare_image_view()

        self.current_compare_name = "unnamed"

        self.append_log("Define Compare Data 완료 (현재 스캔을 비교 기준으로 설정).")
        self.update_tolerance_display()
        self._sync_top_labels()

    # ===== NEW: define된 비교데이터(compare_*)를 별도 폴더로 저장 =====
    def on_save_compare_data_clicked(self):
        if (
            self.compare_pointcloud is None
            and self.compare_depth_array is None
            and self.compare_color_qimage is None
        ):
            self.append_log("[ERROR] 저장할 비교 데이터가 없습니다. 먼저 Define Compare Data를 수행하세요.")
            return

        dlg = CompareSaveDialog(self)
        if dlg.exec() != QDialog.Accepted:
            self.append_log("[INFO] 비교 데이터 저장 취소.")
            return

        folder_name = dlg.final_folder_name
        if not folder_name:
            self.append_log("[ERROR] 폴더명이 생성되지 않아 저장을 중단합니다.")
            return

        try:
            self.save_compare_bundle(folder_name)
            self.append_log(f"[INFO] 비교 데이터 저장 완료: {os.path.join(COMPARE_DIR, folder_name)}")

            self.current_compare_name = folder_name  # 저장된 비교데이터 이름 = 폴더명
            
        except Exception as e:
            self.append_log(f"[ERROR] 비교 데이터 저장 실패: {e}")
        

    # ===== NEW: 저장된 비교데이터 불러오기 =====
    def on_load_compare_data_clicked(self):
        dlg = CompareLoadDialog(self)
        if dlg.exec() != QDialog.Accepted:
            self.append_log("[INFO] 비교 데이터 불러오기 취소.")
            return
        folder = dlg.selected_folder
        if not folder:
            self.append_log("[ERROR] 불러올 폴더가 선택되지 않았습니다.")
            return

        try:
            self.load_compare_bundle(folder)
            self.append_log(f"[INFO] 비교 데이터 불러오기 완료: {os.path.join(COMPARE_DIR, folder)}")

            # 화면 갱신
            if self.btn_point.isChecked():
                self.update_compare_pointcloud_view()
                if self.show_diff_overlay and self.last_pointcloud is not None:
                    self.update_recent_pointcloud_diff_view()
            elif self.btn_depth.isChecked():
                self.update_compare_depth_view()
                if self.last_depth_qimage is not None:
                    self.update_recent_depth_view()
            elif self.btn_image.isChecked():
                self.update_compare_image_view()
                if self.last_color_qimage is not None:
                    self.update_recent_image_view()

            self.update_tolerance_display()

        except Exception as e:
            self.append_log(f"[ERROR] 비교 데이터 불러오기 실패: {e}")

        self._sync_top_labels()

    def save_compare_bundle(self, folder_name: str):
        """comparedata\\folder_name\\folder_name.(png/tiff/ply)로 저장."""
        os.makedirs(COMPARE_DIR, exist_ok=True)

        folder_name = sanitize_folder_component(folder_name)
        save_dir = os.path.join(COMPARE_DIR, folder_name)

        # 덮어쓰기는 다이얼로그에서 이미 결정됨 → 여기서는 그냥 생성
        os.makedirs(save_dir, exist_ok=True)

        base = folder_name

        # --- NEW: 허용오차 설정도 함께 저장(meta.json) ---
        meta_path = os.path.join(save_dir, "meta.json")
        meta = {
            "tol_distance_mm": float(self.get_threshold_mm()),
            "tol_ratio_percent": float(self.get_allow_ratio() * 100.0),
            "gain_db": float(self._get_gain_db()),
            "exposure_ms": float(self._get_exposure_ms()),
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "bundle_name": base,
        }
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta, f, ensure_ascii=False, indent=2)
        self.append_log(f"[COMPARE] META 저장: {meta_path}")

        # PNG
        if self.compare_color_qimage is not None:
            png_path = os.path.join(save_dir, f"{base}.png")
            ok = self.compare_color_qimage.save(png_path, "PNG")
            if ok:
                self.append_log(f"[COMPARE] PNG 저장: {png_path}")
            else:
                raise RuntimeError("COMPARE PNG 저장 실패")
        else:
            self.append_log("[COMPARE] PNG 저장 생략: 2D 이미지 없음")

        # TIFF
        if self.compare_depth_array is not None:
            tiff_path = os.path.join(save_dir, f"{base}.tiff")
            saved = False
            depth = self.compare_depth_array.astype(np.float32)
            if Image is not None:
                img = Image.fromarray(depth)
                img.save(tiff_path, format="TIFF")
                saved = True
            elif cv2 is not None:
                cv2.imwrite(tiff_path, depth)
                saved = True

            if saved:
                self.append_log(f"[COMPARE] TIFF 저장: {tiff_path}")
            else:
                raise RuntimeError("COMPARE TIFF 저장 실패 (PIL/OpenCV 없음)")
        else:
            self.append_log("[COMPARE] TIFF 저장 생략: depth 없음")

        # PLY
        if self.compare_pointcloud is not None and self.compare_pointcloud.size > 0:
            ply_path = os.path.join(save_dir, f"{base}.ply")
            pts = self.compare_pointcloud
            cols = self.compare_colors

            if o3d is not None:
                pcd = o3d.geometry.PointCloud()
                pcd.points = o3d.utility.Vector3dVector(pts)
                if cols is not None:
                    cols_to_use = cols.astype(np.float32)
                    if cols_to_use.max() > 1.0:
                        cols_to_use = cols_to_use / 255.0
                    pcd.colors = o3d.utility.Vector3dVector(cols_to_use)
                o3d.io.write_point_cloud(ply_path, pcd)
            else:
                n = pts.shape[0]
                has_color = cols is not None and cols.shape[0] == n
                with open(ply_path, "w", encoding="utf-8") as f:
                    f.write("ply\n")
                    f.write("format ascii 1.0\n")
                    f.write(f"element vertex {n}\n")
                    f.write("property float x\n")
                    f.write("property float y\n")
                    f.write("property float z\n")
                    if has_color:
                        f.write("property uchar red\n")
                        f.write("property uchar green\n")
                        f.write("property uchar blue\n")
                    f.write("end_header\n")
                    for i in range(n):
                        x, y, z = pts[i]
                        if has_color:
                            c = cols[i].astype(np.float32)
                            if c.max() <= 1.0:
                                c = c * 255.0
                            r, g, b = [int(np.clip(v, 0, 255)) for v in c]
                            f.write(f"{x} {y} {z} {r} {g} {b}\n")
                        else:
                            f.write(f"{x} {y} {z}\n")
            self.append_log(f"[COMPARE] PLY 저장: {ply_path}")
        else:
            self.append_log("[COMPARE] PLY 저장 생략: pointcloud 없음")

    def load_compare_bundle(self, folder_name: str):
        """comparedata\\folder_name\\folder_name.(png/tiff/ply)에서 로드."""
        folder_name = sanitize_folder_component(folder_name)
        load_dir = os.path.join(COMPARE_DIR, folder_name)
        if not os.path.isdir(load_dir):
            raise FileNotFoundError(f"폴더가 없습니다: {load_dir}")

        base = folder_name
        meta_path = os.path.join(load_dir, "meta.json")
        png_path = os.path.join(load_dir, f"{base}.png")
        tiff_path = os.path.join(load_dir, f"{base}.tiff")
        ply_path = os.path.join(load_dir, f"{base}.ply")

        # PNG → compare_color_qimage
        self.compare_color_qimage = None
        if os.path.isfile(png_path):
            qimg = None
            if Image is not None:
                try:
                    img = Image.open(png_path).convert("RGB")
                    arr = np.array(img)
                    h, w, _ = arr.shape
                    qimg = QImage(arr.data, w, h, 3 * w, QImage.Format_RGB888).copy()
                except Exception:
                    qimg = None
            if qimg is None and cv2 is not None:
                img = cv2.imread(png_path, cv2.IMREAD_COLOR)
                if img is not None:
                    img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                    h, w, _ = img.shape
                    qimg = QImage(img.data, w, h, 3 * w, QImage.Format_RGB888).copy()
            self.compare_color_qimage = qimg

        # TIFF → compare_depth_array / compare_depth_qimage(컬러맵)
        self.compare_depth_array = None
        self.compare_depth_qimage = None
        if os.path.isfile(tiff_path):
            depth_np = None
            if cv2 is not None:
                d = cv2.imread(tiff_path, cv2.IMREAD_UNCHANGED)
                if d is not None:
                    depth_np = d.astype(np.float32)
            if depth_np is None and Image is not None:
                img = Image.open(tiff_path)
                depth_np = np.array(img).astype(np.float32)

            if depth_np is not None:
                self.compare_depth_array = depth_np
                rgb_uint8 = depth_to_color_image(depth_np)
                rgb_uint8 = np.ascontiguousarray(rgb_uint8)
                h, w, _ = rgb_uint8.shape
                self.compare_depth_qimage = QImage(
                    rgb_uint8.data, w, h, 3 * w, QImage.Format_RGB888
                ).copy()

        # PLY → compare_pointcloud / compare_colors
        self.compare_pointcloud = None
        self.compare_colors = None
        if os.path.isfile(ply_path):
            if o3d is not None:
                pcd = o3d.io.read_point_cloud(ply_path)
                pts = np.asarray(pcd.points)
                if pts.ndim == 2 and pts.shape[1] == 3:
                    mask = np.isfinite(pts).all(axis=1)
                    pts = pts[mask]
                    self.compare_pointcloud = pts.astype(np.float32)
                    if pcd.has_colors():
                        cols = np.asarray(pcd.colors)
                        cols = cols[mask]
                        self.compare_colors = cols.astype(np.float32)
                    else:
                        self.compare_colors = None
            else:
                # open3d 없으면 최소: point만 읽는 간단 파서(ASCII PLY)
                pts = []
                cols = []
                with open(ply_path, "r", encoding="utf-8", errors="ignore") as f:
                    header = True
                    has_color = False
                    for line in f:
                        line = line.strip()
                        if header:
                            if line.startswith("property uchar red"):
                                has_color = True
                            if line == "end_header":
                                header = False
                            continue
                        parts = line.split()
                        if len(parts) >= 3:
                            x, y, z = map(float, parts[:3])
                            pts.append([x, y, z])
                            if has_color and len(parts) >= 6:
                                r, g, b = map(int, parts[3:6])
                                cols.append([r / 255.0, g / 255.0, b / 255.0])
                if pts:
                    self.compare_pointcloud = np.array(pts, dtype=np.float32)
                    self.compare_colors = np.array(cols, dtype=np.float32) if cols else None

        # --- NEW: 허용오차 설정 불러오기(meta.json) ---
        if os.path.isfile(meta_path):
            try:
                with open(meta_path, "r", encoding="utf-8") as f:
                    meta = json.load(f)

                dist_mm = meta.get("tol_distance_mm", None)
                ratio_pct = meta.get("tol_ratio_percent", None)
                gain_db = meta.get("gain_db", None)
                exposure_ms = meta.get("exposure_ms", None)

                if dist_mm is not None:
                    self.edit_tol_distance.setText(str(dist_mm))
                if ratio_pct is not None:
                    self.edit_tol_ratio.setText(str(ratio_pct))

                # ✅ 게인/노출 입력칸 자동 기입
                if gain_db is not None:
                    self.edit_gain_db.setText(str(gain_db))
                if exposure_ms is not None:
                    self.edit_exposure_ms.setText(str(exposure_ms))

                self.append_log("[COMPARE] META 로드 → 허용오차/카메라 파라미터 자동 Apply")

                # ✅ 허용오차 Apply 자동 실행
                self.on_apply_tolerance_clicked()

                # ✅ 카메라 파라미터 Apply 자동 실행 (실카메라면 적용, 샘플모드면 스킵 로그)
                self.on_apply_camera_params_clicked()

            except Exception as e:
                self.append_log(f"[WARN] meta.json 로드 실패: {e}")

            self.current_compare_name = folder_name  # 저장된 비교데이터 이름 = 폴더명

        else:
            self.append_log("[INFO] meta.json 없음(현재 설정 유지)")

    def on_reset_compare_data_clicked(self):
        # 정의된 비교 데이터(좌측)를 완전히 비움
        self.compare_pointcloud = None
        self.compare_colors = None
        self.compare_depth_qimage = None
        self.compare_depth_array = None
        self.compare_color_qimage = None
        self.current_compare_name = None

        self.append_log("[INFO] 비교 데이터(Define Compare Data)가 리셋되었습니다.")

        # 좌측 뷰어 화면 갱신
        if self.btn_point.isChecked():
            self.compare_viewer.label.setPixmap(QPixmap())
            self.compare_viewer.label.setText("스캔 데이터 출력")
            self.compare_viewer.label.setAlignment(Qt.AlignCenter)
            self.compare_viewer.label.show()
            if self.compare_viewer.gl_view is not None:
                self.compare_viewer.gl_view.hide()

            # 우측은 diff가 켜져있어도 비교 기준이 없으니 일반 뷰로
            if self.last_pointcloud is not None:
                self.update_recent_pointcloud_view()

        elif self.btn_depth.isChecked():
            self.compare_viewer.label.setPixmap(QPixmap())
            self.compare_viewer.label.setText("스캔 데이터 출력")
            self.compare_viewer.label.setAlignment(Qt.AlignCenter)
            self.compare_viewer.label.show()
            if self.last_depth_qimage is not None:
                self.update_recent_depth_view()

        elif self.btn_image.isChecked():
            self.compare_viewer.label.setPixmap(QPixmap())
            self.compare_viewer.label.setText("스캔 데이터 출력")
            self.compare_viewer.label.setAlignment(Qt.AlignCenter)
            self.compare_viewer.label.show()
            if self.last_color_qimage is not None:
                self.update_recent_image_view()

        # 허용오차 표시도 비교 기준이 없으니 '-'로 내려가게
        self.update_tolerance_display()
        self._sync_top_labels()

    # ===== 실제 캡쳐 동작 =====
    def capture_both_depth_and_point(self):
        """
        2D 이미지 + Depth map + Point cloud 를 한 번에 받아서
        '최근 데이터' 변수들(last_*)에 저장.
        """
        if USE_REAL_CAMERA:
            frame2d_and_3d = Frame2DAnd3D()
            show_error(self.camera.capture_2d_and_3d_with_normal(frame2d_and_3d))

            # --- 2D 컬러 이미지 ---
            try:
                frame_2d = frame2d_and_3d.frame_2d()
                color_img = frame_2d.get_color_image()
                color_np = color_img.data().copy()

                if color_np.ndim == 3:
                    # BGR 또는 BGRA → RGB
                    if color_np.shape[2] == 4:
                        bgr = color_np[..., :3]
                    else:
                        bgr = color_np
                    rgb = bgr[..., ::-1].copy()
                    h, w, _ = rgb.shape
                    qimg_color = QImage(
                        rgb.data, w, h, 3 * w, QImage.Format_RGB888
                    ).copy()
                    self.last_color_qimage = qimg_color
                else:
                    self.last_color_qimage = None
            except Exception as e:
                self.last_color_qimage = None
                self.append_log(f"[ERROR] 2D 컬러 이미지 변환 실패: {e}")

            # --- Depth & Point ---
            frame_3d = frame2d_and_3d.frame_3d()

            depth_map = frame_3d.get_depth_map()
            depth_np = depth_map.data().copy().astype(np.float32)
            self.last_depth_array = depth_np

            rgb_uint8 = depth_to_color_image(depth_np)

            mask_valid = np.isfinite(depth_np) & (depth_np > 0)
            if np.any(mask_valid):
                ys, xs = np.where(mask_valid)
                y0, y1 = ys.min(), ys.max() + 1
                x0, x1 = xs.min(), xs.max() + 1

                pad_y = int(0.05 * depth_np.shape[0])
                pad_x = int(0.05 * depth_np.shape[1])

                y0 = max(0, y0 - pad_y)
                y1 = min(depth_np.shape[0], y1 + pad_y)
                x0 = max(0, x0 - pad_x)
                x1 = min(depth_np.shape[1], x1 + pad_x)

                rgb_uint8 = rgb_uint8[y0:y1, x0:x1]

            rgb_uint8 = np.ascontiguousarray(rgb_uint8)
            h, w, _ = rgb_uint8.shape
            qimg = QImage(rgb_uint8.data, w, h, 3 * w, QImage.Format_RGB888).copy()
            self.last_depth_qimage = qimg

            pc = frame_3d.get_untextured_point_cloud()
            pc_np = pc.data().copy()

            pts = pc_np
            if pts.ndim == 3:
                pts = pts.reshape(-1, 3)

            mask = np.isfinite(pts).all(axis=1)
            mask &= (pts[:, 2] > 0)
            pts = pts[mask]

            self.last_pointcloud = pts
            self.last_colors = None

        else:
            # === 샘플 파일 모드 ===

            # 3D 포인트 클라우드
            if o3d is not None:
                pcd_raw = o3d.io.read_point_cloud(SAMPLE_PLY_PATH)
                pts_np = np.asarray(pcd_raw.points)
                mask = np.isfinite(pts_np).all(axis=1)
                pts = pts_np[mask]
                self.last_pointcloud = pts

                if pcd_raw.has_colors():
                    cols_np = np.asarray(pcd_raw.colors)
                    self.last_colors = cols_np[mask]
                else:
                    self.last_colors = None
            else:
                self.last_pointcloud = None
                self.last_colors = None

            # Depth TIFF
            depth_np = None
            if cv2 is not None:
                d = cv2.imread(SAMPLE_TIFF_PATH, cv2.IMREAD_UNCHANGED)
                if d is not None:
                    depth_np = d.astype(np.float32)
            if depth_np is None and Image is not None:
                try:
                    img = Image.open(SAMPLE_TIFF_PATH)
                    depth_np = np.array(img).astype(np.float32)
                except Exception:
                    depth_np = None

            if depth_np is not None:
                self.last_depth_array = depth_np
                rgb_uint8 = depth_to_color_image(depth_np)
                h, w, _ = rgb_uint8.shape
                qimg = QImage(
                    rgb_uint8.data, w, h, 3 * w,
                    QImage.Format_RGB888
                ).copy()
                self.last_depth_qimage = qimg
            else:
                self.last_depth_array = None
                self.last_depth_qimage = None

            # 샘플 2D 이미지
            self.last_color_qimage = None
            if Image is not None:
                try:
                    img = Image.open(SAMPLE_IMG_PATH).convert("RGB")
                    arr = np.array(img)
                    h, w, _ = arr.shape
                    qimg_color = QImage(
                        arr.data, w, h, 3 * w, QImage.Format_RGB888
                    ).copy()
                    self.last_color_qimage = qimg_color
                except Exception:
                    self.last_color_qimage = None

            if self.last_color_qimage is None and cv2 is not None:
                try:
                    img = cv2.imread(SAMPLE_IMG_PATH, cv2.IMREAD_COLOR)
                    if img is not None:
                        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                        h, w, _ = img.shape
                        qimg_color = QImage(
                            img.data, w, h, 3 * w, QImage.Format_RGB888
                        ).copy()
                        self.last_color_qimage = qimg_color
                except Exception:
                    self.last_color_qimage = None

            if self.last_color_qimage is None:
                self.append_log("[WARN] 샘플 2D 이미지 로드 실패: sampleimg.png 확인 필요")

    # ===== 포인트 클라우드 뷰 갱신 =====
    def _update_viewer_with_pointcloud(self, viewer: ScanViewer,
                                       pts: np.ndarray | None,
                                       cols: np.ndarray | None):
        if pts is None or pts.size == 0:
            viewer.label.setText("유효 포인트 없음")
            viewer.label.show()
            return

        pts_vis = pts
        cols_vis = cols

        if o3d is not None:
            pcd = o3d.geometry.PointCloud()
            pcd.points = o3d.utility.Vector3dVector(pts)
            if cols is not None:
                pcd.colors = o3d.utility.Vector3dVector(cols)

            center = pcd.get_center()
            pcd.translate(-center)
            base_R = o3d.geometry.get_rotation_matrix_from_xyz((np.pi, 0, 0))
            pcd.rotate(base_R, center=(0, 0, 0))

            pts_vis = np.asarray(pcd.points)
            if pcd.has_colors():
                cols_vis = np.asarray(pcd.colors)
            else:
                cols_vis = None

        if cols_vis is None and pts_vis is not None and pts_vis.size > 0:
            z = pts_vis[:, 2].astype(np.float32)
            mask = np.isfinite(z)
            if np.any(mask):
                z_valid = z[mask]
                z_min = float(z_valid.min())
                z_max = float(z_valid.max())
                denom = (z_max - z_min) + 1e-6

                norm = (z - z_min) / denom
                norm = np.clip(norm, 0.0, 1.0)

                h = (1 - norm) * (2.0 / 3.0)
                s = np.ones_like(h, dtype=np.float32)
                v = np.ones_like(h, dtype=np.float32)

                r, g, b = hsv_to_rgb_np(h, s, v)
                cols_vis = np.stack([r, g, b], axis=-1).astype(np.float32)
            else:
                cols_vis = None

        viewer.show_pointcloud(pts_vis, cols_vis)

    def update_recent_pointcloud_view(self):
        self._update_viewer_with_pointcloud(
            self.scan_viewer, self.last_pointcloud, self.last_colors
        )

    def update_compare_pointcloud_view(self):
        self._update_viewer_with_pointcloud(
            self.compare_viewer, self.compare_pointcloud, self.compare_colors
        )

    # ===== QImage 뷰 갱신 공통 =====
    def _update_viewer_with_qimage(self, viewer: ScanViewer, qimg: QImage | None, empty_text: str):
        viewer.show_qimage(qimg, empty_text)

    # --- 2D 이미지 ---
    def update_recent_image_view(self):
        self._update_viewer_with_qimage(self.scan_viewer,
                                        self.last_color_qimage,
                                        "이미지 없음")

    def update_compare_image_view(self):
        self._update_viewer_with_qimage(self.compare_viewer,
                                        self.compare_color_qimage,
                                        "이미지 없음")

    # --- 뎁스 ---
    def update_recent_depth_view(self):
        if (
            self.show_diff_overlay and
            self.compare_depth_array is not None and
            self.last_depth_array is not None
        ):
            qimg_diff = self.make_diff_depth_qimage()
            if qimg_diff is not None:
                self._update_viewer_with_qimage(self.scan_viewer,
                                                qimg_diff,
                                                "뎁스 데이터 없음")
                return

        self._update_viewer_with_qimage(self.scan_viewer,
                                        self.last_depth_qimage,
                                        "뎁스 데이터 없음")

    def update_compare_depth_view(self):
        self._update_viewer_with_qimage(self.compare_viewer,
                                        self.compare_depth_qimage,
                                        "뎁스 데이터 없음")

    def update_recent_pointcloud_diff_view(self):
        res = self.make_pointcloud_diff_for_view()
        if res is None:
            self.append_log("[WARN] Point diff 뷰 생성 실패 - 일반 포인트 표시로 대체합니다.")
            self.update_recent_pointcloud_view()
            return

        pts_aligned, colors = res
        self._update_viewer_with_pointcloud(
            self.scan_viewer, pts_aligned, colors
        )

    def compute_depth_diff_stats(self,
                                 ref_depth: np.ndarray,
                                 cur_depth: np.ndarray):
        if ref_depth is None or cur_depth is None:
            return None

        d_ref = ref_depth.astype(np.float32)
        d_cur = cur_depth.astype(np.float32)

        h = min(d_ref.shape[0], d_cur.shape[0])
        w = min(d_ref.shape[1], d_cur.shape[1])
        d_ref = d_ref[:h, :w]
        d_cur = d_cur[:h, :w]

        valid_ref = np.isfinite(d_ref) & (d_ref > 0)
        valid_cur = np.isfinite(d_cur) & (d_cur > 0)
        valid_any = valid_ref | valid_cur

        if not np.any(valid_any):
            return None

        both_valid = valid_ref & valid_cur

        diff_mm = np.zeros_like(d_ref, dtype=np.float32)
        diff_mm[both_valid] = np.abs(d_cur[both_valid] - d_ref[both_valid])

        threshold_mm = self.get_threshold_mm()
        only_one_valid = valid_any & (~both_valid)
        diff_mm[only_one_valid] = threshold_mm

        valid_vals = diff_mm[valid_any]
        mean_mm = float(valid_vals.mean())
        max_mm = float(valid_vals.max())

        diff_ratio = float(np.mean(valid_vals >= threshold_mm))

        mean_cm = mean_mm / 10.0
        max_cm = max_mm / 10.0

        return mean_cm, max_cm, diff_ratio, threshold_mm

    def compute_pointcloud_diff_stats(self,
                                      ref_pts: np.ndarray | None,
                                      cur_pts: np.ndarray | None):
        if o3d is None:
            self.append_log("[WARN] open3d 미설치 - 포인트 클라우드 비교를 건너뜁니다.")
            return None

        if ref_pts is None or cur_pts is None:
            return None

        ref = np.asarray(ref_pts, dtype=np.float32)
        cur = np.asarray(cur_pts, dtype=np.float32)
        if ref.size == 0 or cur.size == 0:
            return None

        pcd_ref = o3d.geometry.PointCloud()
        pcd_ref.points = o3d.utility.Vector3dVector(ref)

        pcd_cur = o3d.geometry.PointCloud()
        pcd_cur.points = o3d.utility.Vector3dVector(cur)

        voxel_size = 2.0
        if voxel_size > 0:
            pcd_ref = pcd_ref.voxel_down_sample(voxel_size)
            pcd_cur = pcd_cur.voxel_down_sample(voxel_size)

        if len(pcd_ref.points) == 0 or len(pcd_cur.points) == 0:
            return None

        align_dist_mm = 30.0
        reg = o3d.pipelines.registration.registration_icp(
            pcd_cur,
            pcd_ref,
            align_dist_mm,
            np.eye(4),
            o3d.pipelines.registration.TransformationEstimationPointToPoint()
        )
        T = reg.transformation

        self.append_log(
            "[DEBUG] ICP 정합: fitness={:.3f}, rmse={:.3f}, 이동={:.2f}mm".format(
                reg.fitness,
                reg.inlier_rmse,
                np.linalg.norm(T[:3, 3])
            )
        )

        pcd_cur_aligned = pcd_cur.transform(T)

        d_cur_to_ref = np.asarray(
            pcd_cur_aligned.compute_point_cloud_distance(pcd_ref)
        )
        d_ref_to_cur = np.asarray(
            pcd_ref.compute_point_cloud_distance(pcd_cur_aligned)
        )

        if d_cur_to_ref.size == 0 or d_ref_to_cur.size == 0:
            return None

        dists = np.concatenate([d_cur_to_ref, d_ref_to_cur])

        threshold_mm = self.get_threshold_mm()
        mean_mm = float(dists.mean())
        max_mm = float(dists.max())
        diff_ratio = float(np.mean(dists >= threshold_mm))

        mean_cm = mean_mm / 10.0
        max_cm = max_mm / 10.0

        return mean_cm, max_cm, diff_ratio, threshold_mm

    def make_diff_depth_qimage(self) -> QImage | None:
        if self.compare_depth_array is None or self.last_depth_array is None:
            return None

        d_ref = self.compare_depth_array.astype(np.float32)
        d_cur = self.last_depth_array.astype(np.float32)

        h = min(d_ref.shape[0], d_cur.shape[0])
        w = min(d_ref.shape[1], d_cur.shape[1])
        d_ref = d_ref[:h, :w]
        d_cur = d_cur[:h, :w]

        valid_ref = np.isfinite(d_ref) & (d_ref > 0)
        valid_cur = np.isfinite(d_cur) & (d_cur > 0)

        valid_any = valid_ref | valid_cur
        if not np.any(valid_any):
            self.append_log("[DEBUG] diff: 유효한 depth 픽셀이 없습니다.")
            return None

        both_valid = valid_ref & valid_cur

        diff_mm = np.zeros_like(d_ref, dtype=np.float32)
        diff_mm[both_valid] = np.abs(d_cur[both_valid] - d_ref[both_valid])

        threshold_mm = self.get_threshold_mm()
        only_one_valid = valid_any & (~both_valid)
        diff_mm[only_one_valid] = threshold_mm

        valid_vals = diff_mm[valid_any]
        mean_diff_mm = float(valid_vals.mean())
        max_diff_mm = float(valid_vals.max())

        self.append_log(
            f"[DEBUG] diff 통계: mean={mean_diff_mm:.2f} mm, "
            f"max={max_diff_mm:.2f} mm, thr={threshold_mm:.1f} mm"
        )

        mask_diff = (diff_mm >= threshold_mm) & valid_any

        rgb_uint8 = depth_to_color_image(d_cur)
        mask_valid_obj = np.isfinite(d_cur) & (d_cur > 0)

        if np.any(mask_valid_obj):
            ys, xs = np.where(mask_valid_obj)
            y0, y1 = ys.min(), ys.max() + 1
            x0, x1 = xs.min(), xs.max() + 1

            pad_y = int(0.05 * d_cur.shape[0])
            pad_x = int(0.05 * d_cur.shape[1])

            y0 = max(0, y0 - pad_y)
            y1 = min(d_cur.shape[0], y1 + pad_y)
            x0 = max(0, x0 - pad_x)
            x1 = min(d_cur.shape[1], x1 + pad_x)

            rgb_uint8 = rgb_uint8[y0:y1, x0:x1]
            mask_diff = mask_diff[y0:y1, x0:x1]
            valid_any_crop = valid_any[y0:y1, x0:x1]
        else:
            valid_any_crop = valid_any

        rgb_uint8 = np.ascontiguousarray(rgb_uint8)
        h, w, _ = rgb_uint8.shape

        num_valid = int(valid_any_crop.sum())
        num_diff = int(mask_diff.sum())
        self.append_log(
            f"[DEBUG] diff 픽셀 수: valid={num_valid}, diff>=thr={num_diff}"
        )

        if num_diff == 0:
            qimg = QImage(rgb_uint8.data, w, h, 3 * w, QImage.Format_RGB888).copy()
            return qimg

        rgb_uint8[mask_diff] = [255, 255, 255]
        qimg = QImage(rgb_uint8.data, w, h, 3 * w, QImage.Format_RGB888).copy()
        return qimg

    def make_pointcloud_diff_for_view(self):
        if o3d is None:
            self.append_log("[WARN] open3d 미설치 - 포인트 diff 뷰를 만들 수 없습니다.")
            return None

        if self.compare_pointcloud is None or self.last_pointcloud is None:
            return None

        ref = np.asarray(self.compare_pointcloud, dtype=np.float32)
        cur = np.asarray(self.last_pointcloud, dtype=np.float32)
        if ref.size == 0 or cur.size == 0:
            return None

        pcd_ref = o3d.geometry.PointCloud()
        pcd_ref.points = o3d.utility.Vector3dVector(ref)

        pcd_cur = o3d.geometry.PointCloud()
        pcd_cur.points = o3d.utility.Vector3dVector(cur)

        voxel_size = 2.0
        if voxel_size > 0:
            pcd_ref = pcd_ref.voxel_down_sample(voxel_size)
            pcd_cur = pcd_cur.voxel_down_sample(voxel_size)

        if len(pcd_ref.points) == 0 or len(pcd_cur.points) == 0:
            return None

        align_dist_mm = 30.0
        reg = o3d.pipelines.registration.registration_icp(
            pcd_cur,
            pcd_ref,
            align_dist_mm,
            np.eye(4),
            o3d.pipelines.registration.TransformationEstimationPointToPoint()
        )

        T = reg.transformation
        self.append_log(
            "[DEBUG] (PointDiff) ICP: fitness={:.3f}, rmse={:.3f}, 이동={:.2f}mm".format(
                reg.fitness,
                reg.inlier_rmse,
                np.linalg.norm(T[:3, 3])
            )
        )

        pcd_cur_aligned = pcd_cur.transform(T)

        pts_cur = np.asarray(pcd_cur_aligned.points)
        pts_ref = np.asarray(pcd_ref.points)
        if pts_cur.size == 0 or pts_ref.size == 0:
            return None

        d_cur_to_ref = np.asarray(
            pcd_cur_aligned.compute_point_cloud_distance(pcd_ref)
        )
        d_ref_to_cur = np.asarray(
            pcd_ref.compute_point_cloud_distance(pcd_cur_aligned)
        )

        threshold_mm = self.get_threshold_mm()

        dist_clipped = np.clip(d_cur_to_ref, 0.0, threshold_mm)
        norm = dist_clipped / (threshold_mm + 1e-6)

        h = (1.0 - norm) * (1.0 / 3.0)  # green ~ red
        s = np.ones_like(h, dtype=np.float32)
        v = np.ones_like(h, dtype=np.float32)

        r, g, b = hsv_to_rgb_np(h, s, v)
        colors_cur = np.stack([r, g, b], axis=1).astype(np.float32)

        mask_large_cur = d_cur_to_ref >= threshold_mm
        colors_cur[mask_large_cur] = 1.0

        mask_large_ref = d_ref_to_cur >= threshold_mm
        pts_ref_far = pts_ref[mask_large_ref]

        if pts_ref_far.size > 0:
            colors_ref_far = np.ones_like(pts_ref_far, dtype=np.float32)
            pts_combined = np.vstack([pts_cur, pts_ref_far])
            colors_combined = np.vstack([colors_cur, colors_ref_far])
        else:
            pts_combined = pts_cur
            colors_combined = colors_cur

        return pts_combined, colors_combined

    def get_threshold_mm(self) -> float:
        default = 10.0
        try:
            text = self.edit_tol_distance.text().strip()
        except AttributeError:
            return default

        if text == "":
            return default

        try:
            val = float(text)
            if val <= 0:
                raise ValueError
            return val
        except ValueError:
            self.append_log("[WARN] 오차 거리(mm) 입력이 잘못되어 기본값 10mm를 사용합니다.")
            self.edit_tol_distance.setText("10")
            return default

    def get_allow_ratio(self) -> float:
        default = 0.02
        try:
            text = self.edit_tol_ratio.text().strip()
        except AttributeError:
            return default

        if text == "":
            return default

        try:
            val_percent = float(text)
            if val_percent < 0:
                raise ValueError
            return val_percent / 100.0
        except ValueError:
            self.append_log("[WARN] 오차 비율(%) 입력이 잘못되어 기본값 2%를 사용합니다.")
            self.edit_tol_ratio.setText("2")
            return default

    def on_apply_tolerance_clicked(self):
        thr_mm = self.get_threshold_mm()
        ratio = self.get_allow_ratio()

        self.append_log(
            f"[INFO] 허용 오차 기준 재적용: 거리={thr_mm:.1f}mm, 비율={ratio*100:.2f}%"
        )

        self.update_tolerance_display()

        if self.show_diff_overlay:
            if self.btn_depth.isChecked() and self.last_depth_array is not None:
                self.update_recent_depth_view()

            if (
                self.btn_point.isChecked()
                and self.last_pointcloud is not None
                and self.compare_pointcloud is not None
            ):
                self.update_recent_pointcloud_diff_view()

    def update_tolerance_display(self):
        depth_result = None
        pc_result = None

        if self.compare_depth_array is not None and self.last_depth_array is not None:
            try:
                depth_result = self.compute_depth_diff_stats(
                    self.compare_depth_array, self.last_depth_array
                )
                m_cm, M_cm, ratio, thr_mm = depth_result
                self.append_log(
                    f"[DEBUG] (Depth) 허용오차: mean={m_cm:.2f} cm, "
                    f"max={M_cm:.2f} cm, ratio={ratio*100:.2f}% (thr={thr_mm:.1f} mm)"
                )
            except Exception as e:
                self.append_log(f"[ERROR] Depth 허용 오차 계산 중 오류: {e}")
                depth_result = None

        if self.compare_pointcloud is not None and self.last_pointcloud is not None:
            try:
                pc_result = self.compute_pointcloud_diff_stats(
                    self.compare_pointcloud, self.last_pointcloud
                )
                if pc_result is not None:
                    m_cm, M_cm, ratio, thr_mm = pc_result
                    self.append_log(
                        f"[DEBUG] (Point) 허용오차: mean={m_cm:.2f} cm, "
                        f"max={M_cm:.2f} cm, ratio={ratio*100:.2f}% (thr={thr_mm:.1f} mm)"
                    )
            except Exception as e:
                self.append_log(f"[ERROR] Point 허용 오차 계산 중 오류: {e}")
                pc_result = None

        result = pc_result if pc_result is not None else depth_result

        if result is None:
            self.label_tol_indicator.setText("-")
            self.label_tol_indicator.setStyleSheet(
                "background-color:white; color:black; font-size:32px; font-weight:bold;"
            )
            return

        mean_cm, max_cm, diff_ratio, threshold_mm = result

        thr_mm = self.get_threshold_mm()
        thr_cm = thr_mm / 10.0
        ratio_limit = self.get_allow_ratio()

        if (mean_cm < thr_cm) and (diff_ratio < ratio_limit):
            self.label_tol_indicator.setText("V")
            self.label_tol_indicator.setStyleSheet(
                "background-color:white; color:lime; font-size:36px; font-weight:bold;"
            )
            src = "Point" if pc_result is not None else "Depth"
            self.append_log(
                f"[DEBUG] 허용오차 판정: OK ({src} 기준, thr={thr_mm:.1f}mm, ratio<{ratio_limit*100:.1f}%)"
            )
        else:
            self.label_tol_indicator.setText("X")
            self.label_tol_indicator.setStyleSheet(
                "background-color:white; color:red; font-size:36px; font-weight:bold;"
            )
            src = "Point" if pc_result is not None else "Depth"
            self.append_log(
                f"[DEBUG] 허용오차 판정: NG ({src} 기준, thr={thr_mm:.1f}mm, ratio>={ratio_limit*100:.1f}%)"
            )

    def on_diff_toggle_clicked(self):
        self.show_diff_overlay = self.btn_diff_toggle.isChecked()
        if self.show_diff_overlay:
            self.btn_diff_toggle.setText("Diff 표시 ON")
        else:
            self.btn_diff_toggle.setText("Diff 표시 OFF")

        if self.btn_depth.isChecked() and self.last_depth_qimage is not None:
            self.update_recent_depth_view()
        elif self.btn_point.isChecked() and self.last_pointcloud is not None:
            if self.show_diff_overlay and self.compare_pointcloud is not None:
                self.update_recent_pointcloud_diff_view()
            else:
                self.update_recent_pointcloud_view()

    # ===== 카메라 파라미터(게인/노출) 파싱 =====
    def _get_gain_db(self) -> float:
        default = 12.0
        try:
            t = self.edit_gain_db.text().strip()
            if t == "":
                return default
            v = float(t)
            if v < 0:
                raise ValueError
            # 권장 범위(0~16) 안내용 클램프 (원하면 제거 가능)
            if v > 16:
                self.append_log("[WARN] 게인 값이 16dB를 초과했습니다. 16으로 제한합니다.")
                v = 16.0
                self.edit_gain_db.setText("16")
            return v
        except Exception:
            self.append_log("[WARN] 게인 입력이 잘못되어 기본값 12dB를 사용합니다.")
            self.edit_gain_db.setText("12")
            return default

    def _get_exposure_ms(self) -> float:
        default = 6.0
        try:
            t = self.edit_exposure_ms.text().strip()
            if t == "":
                return default
            v = float(t)
            if v <= 0:
                raise ValueError
            return v
        except Exception:
            self.append_log("[WARN] 노출(ms) 입력이 잘못되어 기본값 6.0ms를 사용합니다.")
            self.edit_exposure_ms.setText("6.0")
            return default

    # ===== apply 버튼 슬롯 =====
    def on_apply_camera_params_clicked(self):
        gain_db = self._get_gain_db()
        exp_ms = self._get_exposure_ms()

        self.append_log(f"[INFO] 카메라 파라미터 적용: gain={gain_db:.2f} dB, exposure={exp_ms:.2f} ms (2D/3D 동일)")

        if not USE_REAL_CAMERA:
            self.append_log("[INFO] 샘플 모드(USE_REAL_CAMERA=False) → 실제 카메라 적용은 생략합니다.")
            return

        if not self.camera_connected or self.camera is None:
            self.append_log("[ERROR] 카메라 미연결 상태입니다. 먼저 Connect 하세요.")
            return

        self.apply_camera_parameters(gain_db, exp_ms)

    def apply_camera_parameters(self, gain_db: float, exposure_ms: float):
        """
        실제 카메라 UserSet에 2D/3D 동일 게인/노출 적용
        - 3D ExposureSequence = [exposure_ms]
        - 2D ExposureMode(Timed) + ExposureTime = exposure_ms
        - 2D Gain = gain_db
        - 3D Gain = gain_db (파라미터 이름 후보로 시도)
        """
        try:
            user_set = self.camera.current_user_set()

            # --- 3D Exposure (Sequence) ---
            err = user_set.set_float_array_value(Scanning3DExposureSequence.name, [float(exposure_ms)])
            show_error(err)
            self.append_log(f"[OK] 3D ExposureSequence = [{exposure_ms}] ms")

            # --- 2D ExposureMode / ExposureTime ---
            try:
                err = user_set.set_enum_value(Scanning2DExposureMode.name, Scanning2DExposureMode.Value_Timed)
                show_error(err)
            except Exception as e:
                self.append_log(f"[WARN] 2D ExposureMode(Timed) 설정 실패(모델/SDK 차이): {e}")

            try:
                err = user_set.set_float_value(Scanning2DExposureTime.name, float(exposure_ms))
                show_error(err)
                self.append_log(f"[OK] 2D ExposureTime = {exposure_ms} ms")
            except Exception as e:
                self.append_log(f"[WARN] 2D ExposureTime 설정 실패: {e}")

            # --- 2D Gain ---
            try:
                err = user_set.set_float_value(Scanning2DGain.name, float(gain_db))
                show_error(err)
                self.append_log(f"[OK] 2D Gain = {gain_db} dB")
            except Exception as e:
                self.append_log(f"[WARN] 2D Gain 적용 실패(모델/파라미터 차이): {e}")

            # --- 3D Gain (후보 이름으로 시도) ---
            candidates = []
            try:
                candidates.append(Scanning3DGain.name)  # 존재하면 우선
            except Exception:
                pass
            candidates += ["Scanning3DGain", "Scan3DGain", "Scan3D_Gain"]

            applied = False
            for pname in candidates:
                try:
                    err = user_set.set_float_value(pname, float(gain_db))
                    show_error(err)
                    self.append_log(f"[OK] 3D Gain = {gain_db} dB (param='{pname}')")
                    applied = True
                    break
                except Exception:
                    continue

            if not applied:
                self.append_log("[WARN] 3D Gain 파라미터를 찾지 못했거나 적용 실패했습니다. (Scanning3DGain/Scan3DGain 확인 필요)")

            # --- 저장 ---
            try:
                show_error(user_set.save_all_parameters_to_device(), "\nSave the current parameter settings to the selected user set.")
                self.append_log("[OK] UserSet 저장 완료")
            except Exception as e:
                self.append_log(f"[WARN] UserSet 저장 실패(권한/모델 차이): {e}")

        except Exception as e:
            self.append_log(f"[ERROR] 카메라 파라미터 적용 중 예외: {e}")

    def _get_compare_name_for_log(self) -> str:
        # 저장된 비교데이터 이름이 있으면 그 이름, 아니면 unnamed
        name = (self.current_compare_name or "").strip()
        if name == "":
            return "unnamed"
        return name

    def _compute_judgement_stats(self):
        """
        현재 비교 기준과 최근 스캔(last_*)을 기준으로
        판정에 사용된 소스(Point 우선, 없으면 Depth) 통계 반환.

        return:
          dict or None
          {
            "src": "Point" or "Depth",
            "min_mm": float,
            "mean_mm": float,
            "max_mm": float,
            "ratio": float,         # 0~1
            "thr_mm": float,
            "limit_ratio": float,   # 0~1 (입력한 허용비율)
            "verdict": "V" or "X"
          }
        """
        thr_mm = float(self.get_threshold_mm())
        limit_ratio = float(self.get_allow_ratio())

        # --- Point 우선 (open3d 필요) ---
        if o3d is not None and self.compare_pointcloud is not None and self.last_pointcloud is not None:
            try:
                ref = np.asarray(self.compare_pointcloud, dtype=np.float32)
                cur = np.asarray(self.last_pointcloud, dtype=np.float32)
                if ref.size > 0 and cur.size > 0:
                    pcd_ref = o3d.geometry.PointCloud()
                    pcd_ref.points = o3d.utility.Vector3dVector(ref)

                    pcd_cur = o3d.geometry.PointCloud()
                    pcd_cur.points = o3d.utility.Vector3dVector(cur)

                    voxel_size = 2.0
                    if voxel_size > 0:
                        pcd_ref = pcd_ref.voxel_down_sample(voxel_size)
                        pcd_cur = pcd_cur.voxel_down_sample(voxel_size)

                    if len(pcd_ref.points) > 0 and len(pcd_cur.points) > 0:
                        align_dist_mm = 30.0
                        reg = o3d.pipelines.registration.registration_icp(
                            pcd_cur, pcd_ref, align_dist_mm, np.eye(4),
                            o3d.pipelines.registration.TransformationEstimationPointToPoint()
                        )
                        T = reg.transformation
                        pcd_cur_aligned = pcd_cur.transform(T)

                        d_cur_to_ref = np.asarray(pcd_cur_aligned.compute_point_cloud_distance(pcd_ref))
                        d_ref_to_cur = np.asarray(pcd_ref.compute_point_cloud_distance(pcd_cur_aligned))
                        if d_cur_to_ref.size > 0 and d_ref_to_cur.size > 0:
                            dists = np.concatenate([d_cur_to_ref, d_ref_to_cur]).astype(np.float32)

                            min_mm = float(dists.min())
                            mean_mm = float(dists.mean())
                            max_mm = float(dists.max())
                            ratio = float(np.mean(dists >= thr_mm))

                            verdict = "V" if (mean_mm < thr_mm) and (ratio < limit_ratio) else "X"
                            return {
                                "src": "Point",
                                "min_mm": min_mm,
                                "mean_mm": mean_mm,
                                "max_mm": max_mm,
                                "ratio": ratio,
                                "thr_mm": thr_mm,
                                "limit_ratio": limit_ratio,
                                "verdict": verdict,
                            }
            except Exception as e:
                self.append_log(f"[WARN] Point 통계 계산 실패 → Depth로 대체: {e}")

        # --- Depth (open3d 없어도 가능) ---
        if self.compare_depth_array is not None and self.last_depth_array is not None:
            try:
                d_ref = self.compare_depth_array.astype(np.float32)
                d_cur = self.last_depth_array.astype(np.float32)

                h = min(d_ref.shape[0], d_cur.shape[0])
                w = min(d_ref.shape[1], d_cur.shape[1])
                d_ref = d_ref[:h, :w]
                d_cur = d_cur[:h, :w]

                valid_ref = np.isfinite(d_ref) & (d_ref > 0)
                valid_cur = np.isfinite(d_cur) & (d_cur > 0)
                valid_any = valid_ref | valid_cur
                if not np.any(valid_any):
                    return None

                both_valid = valid_ref & valid_cur

                diff_mm = np.zeros_like(d_ref, dtype=np.float32)
                diff_mm[both_valid] = np.abs(d_cur[both_valid] - d_ref[both_valid])

                # 한쪽만 유효하면 threshold로 처리(기존 로직 유지)
                only_one_valid = valid_any & (~both_valid)
                diff_mm[only_one_valid] = thr_mm

                vals = diff_mm[valid_any].astype(np.float32)
                min_mm = float(vals.min())
                mean_mm = float(vals.mean())
                max_mm = float(vals.max())
                ratio = float(np.mean(vals >= thr_mm))

                verdict = "V" if (mean_mm < thr_mm) and (ratio < limit_ratio) else "X"
                return {
                    "src": "Depth",
                    "min_mm": min_mm,
                    "mean_mm": mean_mm,
                    "max_mm": max_mm,
                    "ratio": ratio,
                    "thr_mm": thr_mm,
                    "limit_ratio": limit_ratio,
                    "verdict": verdict,
                }
            except Exception as e:
                self.append_log(f"[WARN] Depth 통계 계산 실패: {e}")
                return None

        return None

    def _save_scan_bundle_to(self, root_dir: str, folder_name: str):
        """
        root_dir\\folder_name\\ 에
        folder_name.png / folder_name.tiff / folder_name.ply 저장
        (on_save_files_clicked 로직을 재사용하는 형태)
        """
        os.makedirs(root_dir, exist_ok=True)
        save_dir = os.path.join(root_dir, folder_name)
        os.makedirs(save_dir, exist_ok=True)

        base = folder_name

        # PNG
        if self.last_color_qimage is not None:
            png_path = os.path.join(save_dir, f"{base}.png")
            ok = self.last_color_qimage.save(png_path, "PNG")
            if ok:
                self.append_log(f"[ERROR-SAVE] PNG 저장: {png_path}")
            else:
                self.append_log("[ERROR-SAVE][WARN] PNG 저장 실패")
        else:
            self.append_log("[ERROR-SAVE][WARN] PNG 없음(저장 생략)")

        # TIFF
        if self.last_depth_array is not None:
            tiff_path = os.path.join(save_dir, f"{base}.tiff")
            try:
                saved = False
                depth = self.last_depth_array.astype(np.float32)
                if Image is not None:
                    img = Image.fromarray(depth)
                    img.save(tiff_path, format="TIFF")
                    saved = True
                elif cv2 is not None:
                    cv2.imwrite(tiff_path, depth)
                    saved = True

                if saved:
                    self.append_log(f"[ERROR-SAVE] TIFF 저장: {tiff_path}")
                else:
                    self.append_log("[ERROR-SAVE][WARN] TIFF 저장 실패(PIL/cv2 없음)")
            except Exception as e:
                self.append_log(f"[ERROR-SAVE][WARN] TIFF 저장 예외: {e}")
        else:
            self.append_log("[ERROR-SAVE][WARN] TIFF 없음(저장 생략)")

        # PLY
        if self.last_pointcloud is not None and self.last_pointcloud.size > 0:
            ply_path = os.path.join(save_dir, f"{base}.ply")
            try:
                pts = self.last_pointcloud
                cols = self.last_colors

                if o3d is not None:
                    pcd = o3d.geometry.PointCloud()
                    pcd.points = o3d.utility.Vector3dVector(pts)
                    if cols is not None:
                        cols_to_use = cols.astype(np.float32)
                        if cols_to_use.max() > 1.0:
                            cols_to_use = cols_to_use / 255.0
                        pcd.colors = o3d.utility.Vector3dVector(cols_to_use)
                    o3d.io.write_point_cloud(ply_path, pcd)
                else:
                    n = pts.shape[0]
                    has_color = cols is not None and cols.shape[0] == n
                    with open(ply_path, "w", encoding="utf-8") as f:
                        f.write("ply\n")
                        f.write("format ascii 1.0\n")
                        f.write(f"element vertex {n}\n")
                        f.write("property float x\n")
                        f.write("property float y\n")
                        f.write("property float z\n")
                        if has_color:
                            f.write("property uchar red\n")
                            f.write("property uchar green\n")
                            f.write("property uchar blue\n")
                        f.write("end_header\n")
                        for i in range(n):
                            x, y, z = pts[i]
                            if has_color:
                                c = cols[i].astype(np.float32)
                                if c.max() <= 1.0:
                                    c = c * 255.0
                                r, g, b = [int(np.clip(v, 0, 255)) for v in c]
                                f.write(f"{x} {y} {z} {r} {g} {b}\n")
                            else:
                                f.write(f"{x} {y} {z}\n")

                self.append_log(f"[ERROR-SAVE] PLY 저장: {ply_path}")
            except Exception as e:
                self.append_log(f"[ERROR-SAVE][WARN] PLY 저장 예외: {e}")
        else:
            self.append_log("[ERROR-SAVE][WARN] PLY 없음(저장 생략)")

        self.append_log(f"[ERROR-SAVE] 저장 완료: {save_dir}")

    def _write_error_excel_row(self, dt: datetime, stats: dict):
        """
        .\\error_log\\YYYYMMDD.xlsx
        - 같은 날짜면 같은 파일에 누적
        - 같은 시각(HHMMSS)이 이미 있으면 그 행 갱신
        - openpyxl 없으면 스킵
        """
        if openpyxl is None:
            self.append_log("[WARN] openpyxl 미설치 → 엑셀 로그 저장을 건너뜁니다. (pip install openpyxl)")
            return

        os.makedirs(ERROR_LOG_DIR, exist_ok=True)

        date_str = dt.strftime("%Y%m%d")
        time_str = dt.strftime("%H%M%S")

        xlsx_path = os.path.join(ERROR_LOG_DIR, f"{date_str}.xlsx")

        headers = [
            "년월일", "시분초",
            "사용된 비교데이터",
            "합격/불합격",
            "적용 오차거리(mm)", "적용 오차비율(%)",
            "판정 오차거리 최소(mm)", "판정 오차거리 평균(mm)", "판정 오차거리 최대(mm)",
            "판정 오차비율(%)",
            "판정 소스(Point)"
        ]

        # 파일 열기/생성
        if os.path.isfile(xlsx_path):
            wb = openpyxl.load_workbook(xlsx_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = date_str
            ws.append(headers)

            # 헤더 스타일 + 폭
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 2)
            ws.freeze_panes = "A2"

        compare_name = self._get_compare_name_for_log()
        verdict = stats.get("verdict", "X")
        thr_mm = float(self.get_threshold_mm())
        ratio_limit_pct = float(self.get_allow_ratio() * 100.0)

        min_mm = float(stats["min_mm"])
        mean_mm = float(stats["mean_mm"])
        max_mm = float(stats["max_mm"])
        ratio_pct = float(stats["ratio"] * 100.0)
        src = stats.get("src", "")

        row_values = [
            date_str, time_str,
            compare_name,
            verdict,
            round(thr_mm, 3), round(ratio_limit_pct, 3),
            round(min_mm, 3), round(mean_mm, 3), round(max_mm, 3),
            round(ratio_pct, 3),
            src
        ]

        # 같은 시각(HHMMSS) 행이 있으면 갱신, 없으면 append
        target_row = None
        for r in range(2, ws.max_row + 1):
            v_date = str(ws.cell(r, 1).value or "")
            v_time = str(ws.cell(r, 2).value or "")
            if v_date == date_str and v_time == time_str:
                target_row = r
                break

        if target_row is None:
            ws.append(row_values)
        else:
            for c, v in enumerate(row_values, start=1):
                ws.cell(row=target_row, column=c).value = v

        wb.save(xlsx_path)
        self.append_log(f"[ERROR-LOG] 엑셀 기록 완료: {xlsx_path} (time={time_str}, verdict={verdict})")

    def _handle_error_on_scan_if_needed(self):
        """
        Scan 직후 호출:
        - ✅ V/X 상관없이 엑셀 기록(갱신 포함)
        - ✅ 단, error 폴더 저장은 X일 때만
        """
        if self.last_capture_dt is None or self.last_capture_name is None:
            return

        stats = self._compute_judgement_stats()
        if stats is None:
            return

        dt = self.last_capture_dt

        # ✅ 1) 엑셀은 V/X 모두 기록(갱신)
        self._write_error_excel_row(dt, stats)

        # ✅ 2) error 폴더 저장은 X일 때만
        if stats.get("verdict") == "X":
            folder_name = self.last_capture_name  # YYYYMMDD_HHMMSS
            self._save_scan_bundle_to(ERROR_SAVE_DIR, folder_name)

    def _ensure_today_log_exists(self) -> str | None:
        """오늘자 xlsx가 없으면 헤더 포함 빈 파일 생성 후 경로 반환."""
        if openpyxl is None:
            self.append_log("[WARN] openpyxl 미설치 → 로그 뷰어를 사용할 수 없습니다.")
            return None

        os.makedirs(ERROR_LOG_DIR, exist_ok=True)
        today = datetime.now().strftime("%Y%m%d")
        xlsx_path = os.path.join(ERROR_LOG_DIR, f"{today}.xlsx")

        # 이미 있으면 그대로
        if os.path.isfile(xlsx_path):
            return xlsx_path

        # 없으면 헤더만 있는 빈 파일 생성
        headers = [
            "년월일", "시분초",
            "사용된 비교데이터",
            "합격/불합격",
            "적용 오차거리(mm)", "적용 오차비율(%)",
            "판정 오차거리 최소(mm)", "판정 오차거리 평균(mm)", "판정 오차거리 최대(mm)",
            "판정 오차비율(%)",
            "판정 소스(Point)"
        ]

        wb = Workbook()
        ws = wb.active
        ws.title = today
        ws.append(headers)

        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(str(h)) + 2)
        ws.freeze_panes = "A2"

        wb.save(xlsx_path)
        self.append_log(f"[INFO] 오늘자 로그 파일이 없어 빈 양식 생성: {xlsx_path}")
        return xlsx_path

    def on_watch_today_log_clicked(self):
        path = self._ensure_today_log_exists()
        if not path:
            return
        dlg = ExcelLogViewerDialog(path, self)
        dlg.exec()

    def on_load_other_log_clicked(self):
        dlg_pick = LogFileLoadDialog(self)
        if dlg_pick.exec() != QDialog.Accepted:
            self.append_log("[INFO] 다른 로그 불러오기 취소.")
            return

        path = dlg_pick.selected_xlsx
        if not path:
            self.append_log("[ERROR] 선택된 로그 파일이 없습니다.")
            return

        dlg = ExcelLogViewerDialog(path, self)
        dlg.exec()


    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, "Exit", "프로그램을 종료하시겠습니까?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            event.accept()
        else:
            event.ignore()


def main():
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()